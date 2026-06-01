"""Write the monthly HR Excel report.

Sheets (in tab order):
- Dashboard                     : executive view -- KPIs + 4 charts.
- Employee Summary              : per-employee late/risk/payroll aggregates.
- Daily Attendance              : every classified employee-day row.
- Daily Trend                   : per-day status counts.
- Missing Punches               : check-in days with no check-out (optional).
- Department Summary            : per-department status counts (optional).
- Employee Reconciliation Details: per-ID audit table.
- Employee Master               : every employee + HR audit flags.
- Overtime                      : rows where overtime actually happened.
- Excluded Employees            : policy exclusions (optional).
- Reconciliation                : the high-level employee-count taxonomy.

Detail sheets get bold blue headers, frozen header row, auto-filter,
and auto-fit column widths. Files land in
REPORT_OUTPUT_DIR/YYYY-MM/hr_report_YYYYMMDD_HHMMSS.xlsx.

The Dashboard is deliberately uncluttered: KPIs on the left, four
consistently-sized charts in a 2x2 grid on the right, gridlines hidden,
and the small backing tables placed below the charts so they do not
crowd the visual area.
"""
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.chart.title import Title
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.fill import ColorChoice as DrawingColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
    RichTextProperties,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config import REPORT_OUTPUT_DIR


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
_SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
_CENTER = Alignment(horizontal="center", vertical="center")

# Reporting Period header (rendered at the top of executive sheets).
_PERIOD_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")     # light blue
_PERIOD_TIMESTAMP_FILL = PatternFill("solid", fgColor="F2F2F2")  # light gray
_PERIOD_FONT = Font(bold=True, size=14, color="1F4E79")
_PERIOD_TIMESTAMP_FONT = Font(bold=True, size=11, color="555555")
_PERIOD_BLOCK_ROWS = 3  # period text + generated timestamp + blank spacer


def _format_period_date(value):
    """Render a date-like value as 'YYYY-MM-DD'. None / NaT -> empty."""
    if value is None:
        return ""
    try:
        ts = pd.to_datetime(value)
        if pd.isna(ts):
            return ""
        return ts.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return str(value)


def _has_period(period_start, period_end):
    """True when at least one of the two bounds is non-empty."""
    return _format_period_date(period_start) != "" or \
        _format_period_date(period_end) != ""


def _write_reporting_period_block(ws, period_start, period_end, n_cols):
    """Render the 3-row Reporting Period header at the top of `ws`.

    Layout:
        Row 1: 'Reporting Period: YYYY-MM-DD to YYYY-MM-DD'
               (bold, size 14, centered, light-blue fill).
        Row 2: 'Generated On: YYYY-MM-DD HH:MM AM/PM'
               (bold, size 11, centered, light-gray fill).
        Row 3: blank spacer (a small placeholder value pins ws.max_row
               so subsequent ws.append() lands at row 4).

    Returns the next row index available for downstream content (4).
    Cells are merged across `n_cols` so the banner spans the report
    width regardless of how many data columns follow.
    """
    start_text = _format_period_date(period_start)
    end_text = _format_period_date(period_end)
    if start_text and end_text:
        period_text = f"Reporting Period: {start_text} to {end_text}"
    elif start_text:
        period_text = f"Reporting Period: from {start_text}"
    elif end_text:
        period_text = f"Reporting Period: through {end_text}"
    else:
        period_text = "Reporting Period: full attendance file"

    generated_text = (
        f"Generated On: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"
    )

    width = max(1, n_cols)

    # Row 1: Reporting Period.
    ws.cell(row=1, column=1, value=period_text)
    if width > 1:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=width)
    c1 = ws.cell(row=1, column=1)
    c1.font = _PERIOD_FONT
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.fill = _PERIOD_HEADER_FILL
    ws.row_dimensions[1].height = 26

    # Row 2: Generated timestamp.
    ws.cell(row=2, column=1, value=generated_text)
    if width > 1:
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=width)
    c2 = ws.cell(row=2, column=1)
    c2.font = _PERIOD_TIMESTAMP_FONT
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.fill = _PERIOD_TIMESTAMP_FILL
    ws.row_dimensions[2].height = 18

    # Row 3: blank spacer. We must write SOMETHING (even a single space)
    # so openpyxl counts the row toward ws.max_row -- otherwise the
    # next ws.append() would overwrite this spacer row.
    ws.cell(row=3, column=1, value=" ")
    ws.row_dimensions[3].height = 6

    return _PERIOD_BLOCK_ROWS + 1

# Every chart on the Dashboard uses the same size so the 2x2 grid stays
# aligned and no chart overlaps a neighbour. Bumped from the original
# 13x7 to 16x9 in the polish pass so HR can read tick labels at a
# glance without clicking into the chart.
_CHART_WIDTH = 16
_CHART_HEIGHT = 9


def _sanitize_row(row):
    """Convert pandas NA / NaT / NaN to None so openpyxl serializes the
    cell as empty instead of raising 'Cannot convert <NA> to Excel'."""
    return [None if pd.isna(v) else v for v in row]


def _style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _autosize_columns(ws, min_width=12, max_width=45):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        longest = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            longest = max(longest, len(str(cell.value)))
        width = min(max_width, max(min_width, longest + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _format_employee_id_columns_as_text(
    ws, header_row, data_start, data_end, n_cols,
):
    """Force any 'Employee ID' column to render as plain TEXT.

    Why: Excel infers a numeric type for large integer IDs (4195162)
    and applies a thousand-separator format (4,195,162) by default,
    which is wrong for an identifier column. Employee IDs are codes,
    not measurements -- they must display exactly as stored.

    Behavior:
    - Walks the header row and locates every column whose header
      contains 'Employee ID' (covers 'Employee ID', 'Raw Employee ID',
      'Canonical Employee ID', 'Old Employee ID', 'Current Employee ID').
    - For each data cell in those columns rewrites the value as a bare
      string, strips any thousand-separator commas and surrounding
      whitespace, and pins number_format='@' so Excel keeps the cell as
      TEXT on any subsequent edit.
    """
    if data_end < data_start:
        return
    for col in range(1, n_cols + 1):
        header = ws.cell(row=header_row, column=col).value
        if not isinstance(header, str) or "Employee ID" not in header:
            continue
        for r in range(data_start, data_end + 1):
            cell = ws.cell(row=r, column=col)
            v = cell.value
            if v is None or v == "":
                continue
            if isinstance(v, float) and v.is_integer():
                text = str(int(v))
            elif isinstance(v, (int, float)):
                text = str(v)
            else:
                text = str(v).replace(",", "").strip()
            cell.value = text
            cell.number_format = "@"


def _write_dataframe(ws, df, start_row):
    """Write a DataFrame anchored at start_row. Returns:
        (header_row, data_start_row, data_end_row, next_blank_row, n_cols).
    """
    n_cols = len(df.columns)
    for r_offset, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_offset, value in enumerate(_sanitize_row(row)):
            ws.cell(row=start_row + r_offset, column=c_offset + 1, value=value)
    _style_header_row(ws, row=start_row, n_cols=n_cols)
    data_start = start_row + 1
    data_end = start_row + len(df)
    next_row = data_end + 2  # one blank row of breathing room
    _format_employee_id_columns_as_text(
        ws, header_row=start_row, data_start=data_start,
        data_end=data_end, n_cols=n_cols,
    )
    return start_row, data_start, data_end, next_row, n_cols


def _build_data_sheet(ws, df, period_start=None, period_end=None):
    """Populate a plain data sheet (header + rows + filter + freeze).

    When `period_start` / `period_end` are provided, a 3-row Reporting
    Period banner is rendered at the very top of the sheet and the
    table header / data / freeze pane / auto-filter all shift down
    accordingly.
    """
    show_period = _has_period(period_start, period_end)
    if df is None or df.empty:
        if show_period:
            _write_reporting_period_block(ws, period_start, period_end, n_cols=1)
            ws.cell(row=_PERIOD_BLOCK_ROWS + 1, column=1, value="(no data)")
        else:
            ws.append(["(no data)"])
        return

    n_cols = len(df.columns)
    if show_period:
        header_row = _write_reporting_period_block(
            ws, period_start, period_end, n_cols=n_cols,
        )
    else:
        header_row = 1

    for r_offset, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_offset, value in enumerate(_sanitize_row(row)):
            ws.cell(row=header_row + r_offset, column=c_offset + 1, value=value)

    data_end = header_row + len(df)
    _style_header_row(ws, row=header_row, n_cols=n_cols)
    ws.freeze_panes = f"A{header_row + 1}"
    last_col_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{data_end}"
    _autosize_columns(ws)
    _format_employee_id_columns_as_text(
        ws, header_row=header_row, data_start=header_row + 1,
        data_end=data_end, n_cols=n_cols,
    )


_PAYABLE_OT_HEADER_FILL = PatternFill("solid", fgColor="548235")   # dark green
_PAYABLE_OT_BODY_FILL = PatternFill("solid", fgColor="E2EFDA")      # light green
_PAYABLE_OT_BORDER_GREEN = "548235"
_NOTES_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")          # light blue


def _apply_payable_overtime_styling(
    ws, payable_col_idx, header_row=1, data_end_row=None,
):
    """Highlight the payable-overtime column to match the mockup.

    - Header cell: dark-green fill with the standard bold-white font.
    - Body cells: light-green fill, bold black font, centered.
    - Adds a thin dark-green border on left/right of every body cell
      and on the bottom of the header so the column visually pops
      out from its neighbors (like the green block in the mockup).

    `header_row` is the row of the table header (1 by default, 4 when
    the sheet carries a Reporting Period banner). `data_end_row`
    defaults to `ws.max_row` for backward compat.
    """
    from openpyxl.styles import Border, Side
    if data_end_row is None:
        data_end_row = ws.max_row
    if data_end_row < header_row:
        return
    header_cell = ws.cell(row=header_row, column=payable_col_idx)
    header_cell.fill = _PAYABLE_OT_HEADER_FILL
    # The shared _style_header_row already set white-bold font; keep it.

    side = Side(style="thin", color=_PAYABLE_OT_BORDER_GREEN)
    body_border = Border(left=side, right=side)
    header_border = Border(left=side, right=side, bottom=side)
    header_cell.border = header_border

    for r in range(header_row + 1, data_end_row + 1):
        cell = ws.cell(row=r, column=payable_col_idx)
        cell.fill = _PAYABLE_OT_BODY_FILL
        cell.font = Font(bold=True, color="000000")
        cell.alignment = _CENTER
        cell.border = body_border


def _append_executive_notes_block(ws, n_cols):
    """Append a compact Notes area below the data.

    Phase-1 readability refresh: the long mockup-style note block
    (8 rows of explanatory bullets + badge) was visually dominating
    the bottom of the sheet on a laptop screen and competing with
    the data table for attention. The replacement is a 2-row note
    that points HR at the important highlighted columns and the
    payable-overtime multiplier, then gets out of the way.
    """
    start_row = ws.max_row + 2  # one blank row of breathing room
    span_end = min(8, n_cols)   # cap merge width so it stays compact

    # Single-row title + body.
    title_cell = ws.cell(
        row=start_row, column=1, value="Notes",
    )
    title_cell.font = Font(bold=True, size=11, color="1F4E79")
    title_cell.fill = _NOTES_TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=span_end)
    ws.row_dimensions[start_row].height = 20

    body = (
        "Payable Overtime (Hours) = Actual Overtime × 1.5; per-row minutes "
        "rounded half-up. Highlighted columns are the metrics most material "
        "to payroll review."
    )
    body_cell = ws.cell(row=start_row + 1, column=1, value=body)
    body_cell.font = Font(size=10, color="333333")
    body_cell.alignment = Alignment(wrap_text=True, vertical="top",
                                    indent=1)
    ws.merge_cells(start_row=start_row + 1, start_column=1,
                   end_row=start_row + 1, end_column=span_end)
    ws.row_dimensions[start_row + 1].height = 32


# ---------------------------------------------------------------------------
# Employee Attendance sheet
# ---------------------------------------------------------------------------
# Audit-focused daily summary, one row per (raw Employee ID, Date). Split-
# shift employees get two pairs of Check-In / Check-Out columns so HR can
# review morning vs evening attendance at a glance. Rows that arrived via
# alias mapping or a manual punch correction are visually highlighted.


_EMP_ATTENDANCE_COLS = [
    "Raw Employee ID", "Canonical Employee Name", "Canonical Employee ID",
    "Date", "Weekday",
    "Shift 1 Check-In", "Shift 1 Check-Out",
    "Shift 2 Check-In", "Shift 2 Check-Out",
    "Total Time", "Source / Notes",
]


def _hhmm(time_str):
    """'HH:MM:SS' -> 'HH:MM' for readable display. Empty/NA stays empty."""
    if time_str is None or (isinstance(time_str, float) and pd.isna(time_str)):
        return ""
    text = str(time_str).strip()
    if not text:
        return ""
    if len(text) >= 5 and text[2] == ":":
        return text[:5]
    return text


def _time_str_to_minutes(time_str):
    """'HH:MM:SS' -> int minutes from midnight. Returns None on bad input."""
    try:
        parts = str(time_str).split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except (ValueError, AttributeError, IndexError):
        return None


def _minutes_between(start_str, end_str):
    """Minutes between two HH:MM[:SS] strings. Handles midnight wrap."""
    start = _time_str_to_minutes(start_str)
    end = _time_str_to_minutes(end_str)
    if start is None or end is None:
        return 0
    if end < start:
        end += 24 * 60
    return end - start


def _format_hours_minutes(total_minutes):
    """Render a minutes count as 'HH:MM'."""
    if total_minutes <= 0:
        return ""
    hours, minutes = divmod(int(total_minutes), 60)
    return f"{hours:02d}:{minutes:02d}"


def _shift_split_boundary(intervals):
    """Time-of-day (minutes from midnight) that splits a 2-interval
    schedule into Shift 1 / Shift 2 partitions. Returns None when the
    schedule has a single interval or an overlapping pair.
    """
    if not intervals or len(intervals) < 2:
        return None
    try:
        first_end_h, first_end_m = map(int, intervals[0][1].split(":"))
        second_start_h, second_start_m = map(int, intervals[1][0].split(":"))
    except (ValueError, IndexError, AttributeError):
        return None
    first_end_min = first_end_h * 60 + first_end_m
    second_start_min = second_start_h * 60 + second_start_m
    if second_start_min <= first_end_min:
        return None  # overlapping or back-to-back -- can't split
    return (first_end_min + second_start_min) // 2


def build_employee_attendance_rows(raw_punches, schedules_df):
    """Return the per-(raw Employee ID, Date) daily-summary rows.

    Public so tests can pin the shape without touching openpyxl. The
    rendering side of this is `_build_employee_attendance_sheet`.

    Behavior:
    - Only Check In + Check Out events count. Breaks have their own sheet.
    - Rows are grouped by `original_employee_id` so HR sees the RAW
      source ID (pre-alias); a parallel column carries the canonical
      Name + ID. If alias mapping is not present in `raw_punches`,
      the canonical ID is reused as the raw ID.
    - Split-shift partitioning uses the midpoint between the first
      interval's end and the second interval's start as the boundary.
      Single-interval employees put every punch into Shift 1; rows
      without a parseable schedule also collapse into Shift 1.
    - Shift 1/2 Check-In = earliest CI in that partition;
      Shift 1/2 Check-Out = latest CO in that partition. Sub-second
      duplicates collapse naturally via min/max.
    - Total Time = sum of paired-shift worked minutes ('HH:MM').
    - Source / Notes documents alias remapping and any manual
      correction that contributed to the row.
    """
    if raw_punches is None or raw_punches.empty:
        return pd.DataFrame(columns=_EMP_ATTENDANCE_COLS)

    df = raw_punches.copy()
    if "original_employee_id" not in df.columns:
        df["original_employee_id"] = df["Employee ID"]
    df = df[df["Punch State"].isin(["Check In", "Check Out"])].copy()
    if df.empty:
        return pd.DataFrame(columns=_EMP_ATTENDANCE_COLS)

    # Canonical name lookup keyed by canonical Employee ID.
    canon_name_by_id = {}
    for ceid, sub in df.groupby("Employee ID"):
        names = sub["First Name"].dropna().astype(str).str.strip()
        names = names[names != ""]
        canon_name_by_id[ceid] = names.iloc[0] if not names.empty else ""

    # Schedule intervals (per canonical name) -- reuse the ScheduleLookup
    # to honor EMP-code + NBSP-normalized matching.
    from metrics_calculator import (
        ScheduleLookup,
        _resolve_schedule_label_column,
    )
    label_col = _resolve_schedule_label_column(schedules_df) or "Working Time"
    schedule_lookup = ScheduleLookup(schedules_df, label_column=label_col)
    intervals_by_name = {
        n: schedule_lookup.match(n)["intervals"]
        for n in canon_name_by_id.values() if n
    }

    rows = []
    for (raw_eid, date), grp in df.groupby(
        ["original_employee_id", "Date"], sort=False
    ):
        canonical_eid = grp["Employee ID"].iloc[0]
        canonical_name = canon_name_by_id.get(canonical_eid, "")
        # Chronological CI/CO pairing: walk events in time order and
        # close each open interval when the first Check Out arrives.
        # This robustly handles split-shifts, event-day manual splits,
        # and sub-second duplicates without depending on a static
        # midpoint boundary (which broke when HR's manual split landed
        # before the configured boundary).
        events = sorted(
            (str(t), s) for t, s in zip(
                grp["Punch Time"].astype(str), grp["Punch State"]
            )
        )
        paired_intervals = []
        open_ci = None
        for time_str, state in events:
            if _time_str_to_minutes(time_str) is None:
                continue
            if state == "Check In":
                if open_ci is None:
                    open_ci = time_str
                # else: subsequent Check In while one is already open
                # -- keep the earlier one (sub-second device duplicates).
            elif state == "Check Out":
                if open_ci is not None:
                    paired_intervals.append((open_ci, time_str))
                    open_ci = None

        # Column-placement strategy:
        # - 2+ chronological pairs AND schedule has 2+ intervals
        #   -> chronological order (Shift 1 = first paired, Shift 2 =
        #   second). Supports event-day splits where the boundary
        #   doesn't fall at the schedule's static midpoint.
        # - Otherwise -> boundary-based slot placement. Each event is
        #   bucketed into Shift 1 / Shift 2 by its time-of-day. This
        #   guarantees that unpaired CIs/COs (e.g. when a raw-ID row
        #   contributes only Check-Ins and the canonical Check-Outs
        #   live in a separate raw-ID row) still appear in their
        #   natural slots instead of disappearing from the audit.
        intervals = intervals_by_name.get(canonical_name) or []
        boundary_min = _shift_split_boundary(intervals)
        s1_ci, s1_co, s2_ci, s2_co = "", "", "", ""
        if len(paired_intervals) >= 2 and len(intervals) >= 2:
            s1_ci, s1_co = paired_intervals[0]
            s2_ci, s2_co = paired_intervals[1]
        else:
            s1_cis, s1_cos, s2_cis, s2_cos = [], [], [], []
            for time_str, state in events:
                t_min = _time_str_to_minutes(time_str)
                if t_min is None:
                    continue
                if boundary_min is None or t_min < boundary_min:
                    ci_b, co_b = s1_cis, s1_cos
                else:
                    ci_b, co_b = s2_cis, s2_cos
                if state == "Check In":
                    ci_b.append(time_str)
                elif state == "Check Out":
                    co_b.append(time_str)
            s1_ci = min(s1_cis) if s1_cis else ""
            s1_co = max(s1_cos) if s1_cos else ""
            s2_ci = min(s2_cis) if s2_cis else ""
            s2_co = max(s2_cos) if s2_cos else ""

        total_min = 0
        for ci_str, co_str in ((s1_ci, s1_co), (s2_ci, s2_co)):
            if ci_str and co_str:
                total_min += _minutes_between(ci_str, co_str)

        notes_parts = []
        is_event_split = (
            "correction_action" in grp.columns
            and (grp["correction_action"] == "event_day_split").any()
        )
        if is_event_split:
            notes_parts.append("Event-day split (manual)")
        elif "is_manual_correction" in grp.columns and grp["is_manual_correction"].any():
            notes_parts.append("Manual correction")
        if raw_eid != canonical_eid:
            notes_parts.append(f"Aliased from {raw_eid} -> {canonical_eid}")
        source_label = "; ".join(notes_parts) if notes_parts else "BioTime"

        try:
            weekday = pd.to_datetime(date).strftime("%a")
        except (ValueError, TypeError):
            weekday = ""

        rows.append({
            "Raw Employee ID": raw_eid,
            "Canonical Employee Name": canonical_name,
            "Canonical Employee ID": canonical_eid,
            "Date": str(date),
            "Weekday": weekday,
            "Shift 1 Check-In": _hhmm(s1_ci),
            "Shift 1 Check-Out": _hhmm(s1_co),
            "Shift 2 Check-In": _hhmm(s2_ci),
            "Shift 2 Check-Out": _hhmm(s2_co),
            "Total Time": _format_hours_minutes(total_min),
            "Source / Notes": source_label,
        })

    result = pd.DataFrame(rows, columns=_EMP_ATTENDANCE_COLS)
    if not result.empty:
        result = result.sort_values(
            by=["Canonical Employee Name", "Date", "Raw Employee ID"]
        ).reset_index(drop=True)
    return result


_ATTENDANCE_ALIAS_FILL = PatternFill("solid", fgColor="FFF2CC")    # light gold
_ATTENDANCE_MANUAL_FILL = PatternFill("solid", fgColor="DDEBF7")   # light blue


def _build_employee_attendance_sheet(ws, df, period_start=None, period_end=None):
    """Render the Employee Attendance audit sheet.

    When `period_start` / `period_end` are provided, a 3-row Reporting
    Period banner is rendered above the table.
    """
    show_period = _has_period(period_start, period_end)
    if df is None or df.empty:
        if show_period:
            _write_reporting_period_block(ws, period_start, period_end, n_cols=1)
            ws.cell(row=_PERIOD_BLOCK_ROWS + 1, column=1,
                    value="(no attendance data available)")
        else:
            ws.append(["(no attendance data available)"])
        return

    n_cols = len(df.columns)
    if show_period:
        header_row = _write_reporting_period_block(
            ws, period_start, period_end, n_cols=n_cols,
        )
    else:
        header_row = 1
    for r_offset, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_offset, value in enumerate(_sanitize_row(row)):
            ws.cell(row=header_row + r_offset, column=c_offset + 1, value=value)
    data_end = header_row + len(df)
    _style_header_row(ws, row=header_row, n_cols=n_cols)
    # Freeze the employee/date identification columns + header row so
    # HR can scroll across many days without losing the row anchor.
    ws.freeze_panes = f"F{header_row + 1}"
    last_col_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{data_end}"
    _autosize_columns(ws, min_width=10, max_width=45)
    _format_employee_id_columns_as_text(
        ws, header_row=header_row, data_start=header_row + 1,
        data_end=data_end, n_cols=n_cols,
    )

    # Visual cue: tint rows that came in via alias mapping (gold) or
    # via a manual punch correction (blue). HR's eye lands on these
    # for audit much faster than reading the Source / Notes column.
    source_col = _EMP_ATTENDANCE_COLS.index("Source / Notes") + 1
    for r in range(header_row + 1, data_end + 1):
        value = ws.cell(row=r, column=source_col).value
        if not isinstance(value, str):
            continue
        fill = None
        if "Aliased" in value:
            fill = _ATTENDANCE_ALIAS_FILL
        elif "Manual" in value:
            fill = _ATTENDANCE_MANUAL_FILL
        if fill is not None:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = fill


# ---------------------------------------------------------------------------
# Employee Summary -- readability refresh (Phase 1)
# ---------------------------------------------------------------------------
# Per-column body tints for the 4 "important" metric columns. These are
# the metrics HR/payroll review most closely; tinting them at low
# saturation makes the eye land on them without overpowering the rest
# of the sheet. Column 9 (Payable Overtime) is the only column that
# gets a bolder green + border treatment via _apply_payable_overtime_styling.
_EMP_SUMMARY_IMPORTANT_TINTS = {
    3:  "FCE4E4",   # No of Absence Days     -- soft red
    7:  "FBE5D6",   # Total Late (Hours)     -- soft peach
    10: "FFF2CC",   # Total Early Leave      -- soft yellow
    12: "DDEBF7",   # Break Time (After Policy) -- soft blue
}
# Subtle zebra stripe for non-highlighted columns on even data rows.
_EMP_SUMMARY_ZEBRA_FILL = "F8F8F8"

# Per-column widths. Names get more room; numeric columns stay narrow
# so the wrapped header text breaks at sensible word boundaries and
# all 14 columns fit on a landscape print page.
_EMP_SUMMARY_COLUMN_WIDTHS = {
    1:  14,  # Employee ID
    2:  36,  # First Name -- widened in the polish pass so full
             # employee names (often 4+ words ending with the EMP code)
             # render on one line instead of wrapping mid-name.
    3:  12,  # No of Absence Days
    4:  12,  # No of Permission Days
    5:  12,  # No of Vacation Days
    6:  12,  # No of Secondment Days
    7:  13,  # Total Late (Hours)
    8:  14,  # Total Over Time (Hours) (Actual)
    9:  14,  # Total Over Time (Payable 1.5x) (Hours)
    10: 13,  # Total Early Leave (Hours)
    11: 12,  # Break Time (Hours)
    12: 14,  # Break Time (After Policy)
    13: 12,  # Friday Compensation Days
    14: 24,  # Friday Worked Dates
}


def _apply_employee_summary_body_fills(
    ws, header_row, data_start, data_end, n_cols,
):
    """Apply zebra striping + important-column tints to the data body.

    - Every other data row gets a very light gray fill on the
      non-highlighted columns (subtle, low contrast).
    - The 4 "important" columns (Absence, Late, Early Leave, Break
      After Policy) get a solid soft-color tint on EVERY row so HR's
      eye can find them column-wise.
    - Column 9 (Payable Overtime) is skipped here; its bolder green
      treatment is applied later by `_apply_payable_overtime_styling`
      and would otherwise be overwritten.
    """
    zebra = PatternFill("solid", fgColor=_EMP_SUMMARY_ZEBRA_FILL)
    tint_fills = {
        col: PatternFill("solid", fgColor=color)
        for col, color in _EMP_SUMMARY_IMPORTANT_TINTS.items()
    }
    for r in range(data_start, data_end + 1):
        is_even = ((r - data_start) % 2) == 1
        for c in range(1, n_cols + 1):
            if c == 9:
                continue  # Payable OT styled separately.
            if c in tint_fills:
                ws.cell(row=r, column=c).fill = tint_fills[c]
            elif is_even:
                ws.cell(row=r, column=c).fill = zebra


def _style_employee_summary_header(ws, header_row, n_cols):
    """Style the header row so the long column names wrap cleanly
    over 2-3 lines instead of overflowing into neighboring cells.

    Header font, fill, and centering come from `_style_header_row`;
    this helper then sets wrap_text + a taller row height so the
    multi-line labels are fully visible without clipping.
    """
    _style_header_row(ws, row=header_row, n_cols=n_cols)
    wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, n_cols + 1):
        ws.cell(row=header_row, column=c).alignment = wrap
    ws.row_dimensions[header_row].height = 42


def _apply_employee_summary_print_settings(ws, header_row):
    """Configure the sheet for printer-friendly output.

    - Landscape orientation so the 14 columns fit one one printed page.
    - Fit-to-width = 1 page; fit-to-height = 0 (let it grow vertically).
    - Repeat rows 1..header_row on every printed page so the Reporting
      Period banner + column headers are visible above the data on
      page 2+, not just page 1.
    """
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    # Repeat the banner + header band on every printed page.
    ws.print_title_rows = f"1:{header_row}"
    ws.print_options.horizontalCentered = True


def _build_executive_employee_sheet(ws, df, period_start=None, period_end=None):
    """Render the executive Employee Summary sheet (readability refresh).

    14 columns expected (unchanged in this Phase-1 refresh):
        1.  Employee ID
        2.  First Name
        3.  No of Absence Days       (reduced by Friday compensation)
        4.  No of Permission Days
        5.  No of Vacation Days
        6.  No of Secondment Days
        7.  Total Late (Hours)
        8.  Total Over Time (Hours) (Actual)
        9.  Total Over Time (Payable 1.5x) (Hours)
        10. Total Early Leave (Hours)
        11. Break Time (Hours)
        12. Break Time (After Policy)
        13. Friday Compensation Days  (Gaming Friday Compensation)
        14. Friday Worked Dates       (audit list of paired Fridays)

    Readability features (no calculation changes):
    - Wrapped, taller header row so long labels read on 2-3 lines.
    - Frozen header row AND first two columns so Employee ID + First
      Name + the header stay visible while scrolling.
    - Consistent numeric formats: `#,##0` for integer day counts,
      `0.0` for hour / fractional-day values.
    - Subtle zebra striping for the non-highlighted columns.
    - Soft per-column tints on the 4 metrics most material to
      payroll review (Absence / Late Hours / Early Leave / Break
      After Policy). Column 9 (Payable Overtime) keeps its bolder
      green + bordered treatment.
    - Compact 2-row Notes area instead of the 9-row block.
    - Print-friendly: landscape, fit to width, repeat banner +
      header rows on every printed page.
    """
    show_period = _has_period(period_start, period_end)
    if df is None or df.empty:
        if show_period:
            _write_reporting_period_block(ws, period_start, period_end, n_cols=1)
            ws.cell(row=_PERIOD_BLOCK_ROWS + 1, column=1, value="(no data)")
        else:
            ws.append(["(no data)"])
        return

    n_cols = len(df.columns)
    if show_period:
        header_row = _write_reporting_period_block(
            ws, period_start, period_end, n_cols=n_cols,
        )
    else:
        header_row = 1
    for r_offset, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_offset, value in enumerate(_sanitize_row(row)):
            ws.cell(row=header_row + r_offset, column=c_offset + 1, value=value)
    data_start_row = header_row + 1
    data_end_row = header_row + len(df)

    # Header band with wrapped, taller labels.
    _style_employee_summary_header(ws, header_row=header_row, n_cols=n_cols)

    # Filter spans the header + body.
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(n_cols)}{data_end_row}"
    )

    # Explicit column widths -- skip _autosize_columns so the wrapped
    # header doesn't get expanded into a single ugly long line.
    for col_idx, width in _EMP_SUMMARY_COLUMN_WIDTHS.items():
        if col_idx <= n_cols:
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Integer columns: the three whole-day-count columns (4 Permission,
    # 5 Vacation, 6 Secondment) and the Gaming Friday Compensation
    # Days column (13). Employee ID (col 1) is intentionally NOT
    # included -- it gets its own TEXT (@) format below.
    _format_numeric_cells(
        ws,
        rows=range(data_start_row, data_end_row + 1),
        cols=[4, 5, 6, 13],
        number_format="#,##0",
    )
    # 1-decimal columns: 3 Absence (fractional for split-shift),
    # 7 Late, 8 OT Actual, 9 OT Payable, 10 Early Leave, 11 Break,
    # 12 Break After Policy.
    _format_numeric_cells(
        ws,
        rows=range(data_start_row, data_end_row + 1),
        cols=[3, 7, 8, 9, 10, 11, 12],
        number_format="0.0",
    )

    # Zebra stripes + per-column tints on the 4 important metrics.
    # MUST run before the payable-OT styling so col 9 can overwrite
    # the zebra/tint fill with its prominent green treatment.
    _apply_employee_summary_body_fills(
        ws, header_row=header_row,
        data_start=data_start_row, data_end=data_end_row,
        n_cols=n_cols,
    )

    # Hero column: Payable Overtime (col 9) -- bolder green + border.
    _apply_payable_overtime_styling(
        ws, payable_col_idx=9,
        header_row=header_row, data_end_row=data_end_row,
    )

    # Pin Employee ID column(s) to TEXT (@) so Excel never inserts
    # thousand separators into the identifier. Runs AFTER the numeric
    # formatters above so this is the last write to those cells.
    _format_employee_id_columns_as_text(
        ws, header_row=header_row, data_start=data_start_row,
        data_end=data_end_row, n_cols=n_cols,
    )

    # Freeze the header row AND the first two columns so HR can scroll
    # across many metric columns without losing the row anchor
    # (Employee ID + First Name) or the column anchor (the wrapped
    # header band).
    ws.freeze_panes = f"C{header_row + 1}"

    # Print-friendly: landscape, fit-to-width, repeat banner + header.
    _apply_employee_summary_print_settings(ws, header_row=header_row)

    # Compact Notes area below the data.
    _append_executive_notes_block(ws, n_cols=n_cols)


def _apply_daily_conditional_formatting(ws, df, data_start_row=2):
    """Color-code each row in Daily Attendance by its dominant status.

    Rules are applied in priority order with stopIfTrue=True so each
    row gets exactly one fill, picking the most important condition.
    Priority (highest first): Excluded, Leave, Approved Excuse, Late,
    Early Leave, Missing Check Out, Overtime.

    `data_start_row` is the spreadsheet row of the FIRST data row
    (defaults to 2). It moves down to 5 when the sheet carries a
    Reporting Period banner.
    """
    if df is None or df.empty:
        return

    cols = list(df.columns)
    n_rows = len(df)
    n_cols = len(cols)
    last_col_letter = get_column_letter(n_cols)
    data_end_row = data_start_row + n_rows - 1
    range_str = (
        f"A{data_start_row}:{last_col_letter}{data_end_row}"
    )
    # Conditional-formatting formulas reference the first data row
    # (Excel auto-shifts the row number for subsequent rows in the
    # range). Keep this in lockstep with `data_start_row` so the
    # rules still hit the same cells after the period banner pushed
    # everything down.
    R = data_start_row

    def col_letter(name):
        return get_column_letter(cols.index(name) + 1) if name in cols else None

    is_late_L = col_letter("is_late")
    unexcused_L = col_letter("unexcused_delay_minutes")
    overtime_L = col_letter("overtime_minutes")
    early_leave_L = col_letter("early_leave_minutes")
    early_leave_anomaly_L = col_letter("early_leave_anomaly")
    missing_co_L = col_letter("missing_check_out")
    status_L = col_letter("attendance_status")
    excluded_L = col_letter("is_excluded")

    def _add(formula, fill_hex, font_hex="000000", bold=False):
        ws.conditional_formatting.add(
            range_str,
            FormulaRule(
                formula=[formula],
                fill=PatternFill("solid", fgColor=fill_hex),
                font=Font(bold=bold, color=font_hex),
                stopIfTrue=True,
            ),
        )

    # Priority order (highest first). The first matching rule wins per row.
    if excluded_L:
        _add(f"${excluded_L}{R}=TRUE", "C8A2C8")              # light violet
    if status_L:
        _add(f'${status_L}{R}="Leave"', "BFBFBF")             # gray
        _add(f'${status_L}{R}="Approved Excuse"', "DDEBF7")   # light blue
    # Early-leave ANOMALIES outrank normal Late/Early Leave so HR can
    # spot data-quality issues at a glance (distinct dark red).
    if early_leave_anomaly_L:
        _add(
            f"${early_leave_anomaly_L}{R}=TRUE",
            "8B0000", font_hex="FFFFFF", bold=True,           # dark red + white bold
        )
    if is_late_L and unexcused_L:
        _add(
            f"OR(${is_late_L}{R}=TRUE,${unexcused_L}{R}>0)",
            "C00000", font_hex="FFFFFF", bold=True,           # red + white bold
        )
    if early_leave_L:
        _add(f"${early_leave_L}{R}>0", "CC6600",
             font_hex="FFFFFF")                                # dark orange
    if missing_co_L:
        _add(f"${missing_co_L}{R}=TRUE", "FFFF00", bold=True)  # yellow
    if overtime_L:
        _add(f"${overtime_L}{R}>0", "C6EFCE",
             font_hex="006100")                              # green


# ---------------------------------------------------------------------------
# Chart-polish helpers
# ---------------------------------------------------------------------------
# All Dashboard charts share the same visual language so the eye learns
# the layout once: bold title + smaller italic subtitle + clean axes +
# right-side legend for pies / no-legend for single-series bars and lines.

def _chart_title(title_text, subtitle_text=None):
    """Build a `Title` containing the title and an optional subtitle.

    openpyxl renders a chart `Title` as a single text block by default.
    To get the executive "Title + (subtitle)" look used by Power BI
    dashboards we hand-build a RichText with TWO paragraphs:
      P1 = title (bold, larger)
      P2 = subtitle (italic, smaller, gray)
    """
    # Sizes are in 1/100 pt (so sz=1400 -> 14pt, sz=1000 -> 10pt).
    title_run = RegularTextRun(
        rPr=CharacterProperties(b=True, sz=1400, solidFill="1F4E79"),
        t=title_text,
    )
    title_p = Paragraph(
        pPr=ParagraphProperties(
            defRPr=CharacterProperties(b=True, sz=1400, solidFill="1F4E79"),
        ),
        r=[title_run],
    )
    paragraphs = [title_p]
    if subtitle_text:
        subtitle_run = RegularTextRun(
            rPr=CharacterProperties(
                b=False, i=True, sz=1000, solidFill="595959",
            ),
            t=subtitle_text,
        )
        subtitle_p = Paragraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(
                    b=False, i=True, sz=1000, solidFill="595959",
                ),
            ),
            r=[subtitle_run],
        )
        paragraphs.append(subtitle_p)
    rich = RichText(p=paragraphs)
    # openpyxl's Title.tx accepts a RichText directly via the 'rich'
    # attribute of its Text wrapper. Using the Title(tx=...) helper
    # works across the 3.x line.
    from openpyxl.chart.data_source import NumDataSource  # noqa: F401
    from openpyxl.chart.text import Text as ChartTextWrap
    text = ChartTextWrap()
    text.rich = rich
    return Title(tx=text, overlay=False)


def _axis_label_text(rotation_60000ths=0, size_100ths=900):
    """Return a `RichText` for the axis-label text properties.

    `rotation_60000ths` rotates the labels (negative = clockwise).
    -45 degrees = -2_700_000.
    `size_100ths` sets the font size in 1/100 pt (so 900 -> 9 pt).
    """
    return RichText(
        bodyPr=RichTextProperties(rot=rotation_60000ths),
        p=[Paragraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=size_100ths),
            ),
        )],
    )


def _pie_chart(title, labels, values, subtitle=None):
    """Pie chart with executive defaults.

    Polish:
    - Title + small italic subtitle (built via `_chart_title`).
    - Data labels show ONLY the percentage (no category name) so they
      stay readable inside small slices. The right-side legend
      carries the category names, so there is no information loss.
    - Labels position 'bestFit' lets Excel push tight labels outside
      the slice automatically (with leader lines) so the 78%/19%
      block doesn't crowd the tiny 3%/0% slices.
    - Legend pinned to the right with `overlay=False` so it never
      sits on top of the slices.
    """
    chart = PieChart()
    chart.title = _chart_title(title, subtitle)
    chart.add_data(values, titles_from_data=False)
    chart.set_categories(labels)
    chart.width = _CHART_WIDTH
    chart.height = _CHART_HEIGHT
    chart.dataLabels = DataLabelList(
        showCatName=False,
        showPercent=True,
        showSerName=False,
        showVal=False,
        showLegendKey=False,
        dLblPos="bestFit",
    )
    chart.legend = Legend()
    chart.legend.position = "r"
    chart.legend.overlay = False
    return chart


def _bar_chart(title, labels, values, horizontal=False, subtitle=None,
                x_axis_title=None, y_axis_title=None):
    """Bar chart with executive defaults.

    Polish (for the Dashboard's horizontal bar charts):
    - Title + italic subtitle (e.g. "By Total Late Hours (Hours)") so
      HR can read the metric without clicking the chart.
    - Data labels on EVERY bar, positioned at the bar's outer end
      (`dLblPos="outEnd"`) so the numeric value sits right next to
      the bar tip instead of being hidden inside it. This is the #1
      readability fix per the latest review.
    - Optional X / Y axis titles -- for horizontal bars the X-axis
      carries the metric label ("Late Hours (Hours)") and the Y-axis
      carries the category dimension ("Employee").
    - Axis tick labels rendered at 9pt so employee names and bar
      values are legible on a laptop screen.
    - Single-series bar -- no legend (the title + subtitle already
      identify the metric; a legend would just steal plot area).
    """
    chart = BarChart()
    if horizontal:
        chart.type = "bar"
    chart.title = _chart_title(title, subtitle)
    chart.legend = None
    chart.add_data(values, titles_from_data=False)
    chart.set_categories(labels)
    chart.width = _CHART_WIDTH
    chart.height = _CHART_HEIGHT
    # Numeric value at the outside end of each bar.
    chart.dataLabels = DataLabelList(
        showVal=True,
        showCatName=False,
        showSerName=False,
        showLegendKey=False,
        dLblPos="outEnd",
    )
    if x_axis_title:
        chart.x_axis.title = x_axis_title
    if y_axis_title:
        chart.y_axis.title = y_axis_title
    # Force category labels (employee names) to render on the "low"
    # side -- for a horizontal bar chart this is the left, where HR
    # expects to read them. Without this Excel can flip the labels
    # to the bar's other end when bars become narrow.
    chart.y_axis.tickLblPos = "low"
    chart.x_axis.tickLblPos = "low"
    # Slightly larger axis label font (9pt) for laptop readability.
    chart.x_axis.txPr = _axis_label_text(rotation_60000ths=0, size_100ths=900)
    chart.y_axis.txPr = _axis_label_text(rotation_60000ths=0, size_100ths=900)
    return chart


def _write_chart_annotation(
    ws, top_row, col_left, col_right,
    chart_label, body, accent_primary,
):
    """Render a small "How to read" annotation block adjacent to a chart.

    Two cells stacked vertically, merged across `col_left:col_right`:
        Row N+0 -- "How to read - {chart_label}" on the chart's
                   section accent colour (white bold text).
        Row N+1 -- a short italic description that explains what the
                   chart shows and which direction is "good" vs "bad".
                   Wrapped, max 2 lines on a laptop screen.

    `accent_primary` should be the same primary colour the chart's
    section uses elsewhere (e.g. Attendance Risk red for Late charts,
    Payroll Impact green for Overtime charts) so the eye links the
    annotation back to its chart.
    """
    from openpyxl.utils import column_index_from_string
    left_idx = column_index_from_string(col_left)
    right_idx = column_index_from_string(col_right)

    # Title row: bold white on the section accent.
    ws.merge_cells(
        start_row=top_row, start_column=left_idx,
        end_row=top_row, end_column=right_idx,
    )
    th = ws.cell(row=top_row, column=left_idx)
    th.value = f"How to read - {chart_label}"
    th.font = Font(bold=True, size=10, color="FFFFFF")
    th.fill = PatternFill("solid", fgColor=accent_primary)
    th.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[top_row].height = 18

    # Body row: italic small text on a very light gray strip.
    body_row = top_row + 1
    ws.merge_cells(
        start_row=body_row, start_column=left_idx,
        end_row=body_row, end_column=right_idx,
    )
    bc = ws.cell(row=body_row, column=left_idx)
    bc.value = body
    bc.font = Font(italic=True, size=9, color="333333")
    bc.fill = PatternFill("solid", fgColor="F8F8F8")
    bc.alignment = Alignment(
        horizontal="left", vertical="top",
        wrap_text=True, indent=1,
    )
    ws.row_dimensions[body_row].height = 30


def _line_chart(title, labels, values_with_header,
                 subtitle=None, x_axis_title=None, y_axis_title=None):
    """Line chart with executive defaults.

    Polish:
    - Title + smaller italic subtitle (e.g. "(Number of Late Cases)").
    - Optional X/Y axis titles -- HR managers can see the units
      without clicking into the chart ("Date" / "Late Cases").
    - Major gridlines on the Y-axis make it easy to read off values
      without hovering each point.
    - Each data point gets a circle marker so HR can see day-to-day
      differences even when the line is short / flat.
    - X-axis labels rotated -45deg so long date strings
      (2026-05-31) don't overlap each other on a narrow chart.
    - `varyColors=False` keeps the single-series line one consistent
      colour instead of cycling theme colours per segment.
    """
    chart = LineChart()
    chart.title = _chart_title(title, subtitle)
    chart.legend = None
    chart.add_data(values_with_header, titles_from_data=True)
    chart.set_categories(labels)
    chart.width = _CHART_WIDTH
    chart.height = _CHART_HEIGHT
    chart.varyColors = False
    if x_axis_title:
        chart.x_axis.title = x_axis_title
    if y_axis_title:
        chart.y_axis.title = y_axis_title
    chart.y_axis.majorGridlines = ChartLines()
    # Markers on each data point.
    for s in chart.series:
        s.marker = Marker(symbol="circle", size=7)
        # Make the line slightly thicker so the trend reads at a glance.
        s.graphicalProperties = GraphicalProperties(
            ln=LineProperties(w=22000),  # ~1.75pt in EMU
        )
    # Rotate X-axis labels (-45deg) for date readability.
    chart.x_axis.txPr = _axis_label_text(
        rotation_60000ths=-2_700_000, size_100ths=900,
    )
    return chart


def _format_numeric_cells(ws, rows, cols, number_format):
    """Apply thousands separator + centered alignment to a numeric block."""
    align = Alignment(horizontal="center", vertical="center")
    for row in rows:
        for col in cols:
            cell = ws.cell(row=row, column=col)
            cell.number_format = number_format
            cell.alignment = align


# ---------------------------------------------------------------------------
# Dashboard layout helpers (Phase 1 executive redesign)
# ---------------------------------------------------------------------------
# The Dashboard is laid out in 4 vertical zones for executive readability.
# All visible content lives within columns A:M (13 columns) so a 4-card
# row fits the default Excel viewport without horizontal scrolling.
#
#   ZONE 1 -- Header                  (rows 1-5)
#       Big title at row 1 (merged A:M); Reporting Period + Generated
#       On banner at rows 2-3; spacers at rows 4-5.
#
#   ZONE 2 -- Hero KPI cards          (rows 6-10)
#       Four cards of equal width (3 cols each), tiled A:C / D:F /
#       G:I / J:L. Column M is a small right-side margin. Cards are
#       separated by their colored label bands rather than gutters.
#
#   ZONE 3 -- Charts                  (rows 11-?)
#       Five charts in a 2x2+1 grid. Left column anchored at A,
#       right column anchored at H -- aligns visually with the card
#       boundaries and stays within the A:M visible window.
#
#   ZONE 4 -- Detail KPI sections     (rows 82-?)
#       Two side-by-side panels (A:F and H:M) for analysts auditing
#       the executive figures.
#
# Backing tables that drive the pie/bar charts live on a SEPARATE
# hidden helper sheet ("Dashboard Data") -- see
# `_write_dashboard_backing_tables`. They used to live below row 110
# of this sheet; moving them to a hidden helper keeps the visible
# Dashboard area focused on KPIs + charts + sections.
#
# Constants below are the single source of truth for these zones.
_DASH_CARDS_ROW = 6           # top of the 4 hero KPI cards
_DASH_CARD_HEIGHT = 5         # rows per card
_DASH_CHARTS_ROW = 11         # top-left chart anchor (row)
_DASH_CHART_ROW_STRIDE = 22   # vertical gap between successive chart rows
_DASH_DETAIL_SECTIONS_ROW = 82
_DASH_HELPER_SHEET_NAME = "Dashboard Data"

# Per-section accent palette. Each entry is (primary, soft) for the
# header fill + the soft body fill on the matching detail block.
_DASH_SECTION_COLORS = {
    "Workforce":       ("305496", "DDEBF7"),   # corporate blue
    "Attendance Risk": ("C00000", "F8CBAD"),   # red / peach
    "Payroll Impact":  ("548235", "E2EFDA"),   # green
    "Data Quality":    ("BF8F00", "FFF2CC"),   # gold
}

# Card column anchors. 4 cards x 3 cols each, no inter-card gutter
# (cards are visually separated by their colored label bands). Total
# 12 cols (A:L) -- fits inside the visible A:M window with col M as
# a thin right-side margin.
_DASH_CARD_COLS = [
    ("A", "C"),   # Workforce
    ("D", "F"),   # Attendance Risk
    ("G", "I"),   # Payroll Impact
    ("J", "L"),   # Data Quality
]
# Detail section columns: 2 panels side-by-side, each 6 cols wide,
# with column G as the gutter. Stays within A:M.
_DASH_DETAIL_COLS = [
    ("A", "F"),   # left detail panel
    ("H", "M"),   # right detail panel
]
# Banner / title merges span A:M.
_DASH_BANNER_MERGE = ("A", "M")


def _write_dashboard_title(ws, period_start, period_end):
    """Render the Dashboard's header zone (rows 1-5).

    Row 1: large title (`HR Reporting Dashboard`), 24pt bold centered.
    Row 2-3: Reporting Period banner shared with the other sheets (or
    the legacy "Generated YYYY-MM-DD HH:MM" subtitle when no period
    bounds are provided -- preserves the existing tests).
    Rows 4-5: visual breathing room.

    Returns the row index where downstream content (cards) should
    start. Always 6 -- the title zone is fixed-height by design.
    """
    banner_left, banner_right = _DASH_BANNER_MERGE
    banner_range = lambda row: f"{banner_left}{row}:{banner_right}{row}"

    ws.merge_cells(banner_range(1))
    title = ws[f"{banner_left}1"]
    title.value = "HR Reporting Dashboard"
    title.font = Font(bold=True, size=24, color="1F4E79")
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.fill = PatternFill("solid", fgColor="F2F2F2")
    ws.row_dimensions[1].height = 38

    if _has_period(period_start, period_end):
        start_text = _format_period_date(period_start)
        end_text = _format_period_date(period_end)
        if start_text and end_text:
            period_text = f"Reporting Period: {start_text} to {end_text}"
        elif start_text:
            period_text = f"Reporting Period: from {start_text}"
        elif end_text:
            period_text = f"Reporting Period: through {end_text}"
        else:
            period_text = "Reporting Period: full attendance file"
        ws.merge_cells(banner_range(2))
        c = ws[f"{banner_left}2"]
        c.value = period_text
        c.font = _PERIOD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = _PERIOD_HEADER_FILL
        ws.row_dimensions[2].height = 26

        ws.merge_cells(banner_range(3))
        g = ws[f"{banner_left}3"]
        g.value = (
            f"Generated On: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"
        )
        g.font = _PERIOD_TIMESTAMP_FONT
        g.alignment = Alignment(horizontal="center", vertical="center")
        g.fill = _PERIOD_TIMESTAMP_FILL
        ws.row_dimensions[3].height = 18

        ws.row_dimensions[4].height = 6
    else:
        # Legacy subtitle path -- callers that don't pass period bounds
        # still get a useful "Generated" line at row 2 (matches the
        # contract used by older callers and the matching tests).
        ws.merge_cells(banner_range(2))
        subtitle = ws[f"{banner_left}2"]
        subtitle.value = (
            f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        subtitle.font = Font(italic=True, color="555555", size=10)
        subtitle.alignment = Alignment(horizontal="left", vertical="center")

    return _DASH_CARDS_ROW


def _write_kpi_card(ws, top_row, col_left, col_right,
                     section_label, descriptor, value, number_format,
                     accent_primary, accent_soft):
    """Render one of the 4 hero KPI cards.

    Card layout (5 rows tall, 3 columns wide):
        Row N+0  -- section_label (e.g. "WORKFORCE"), white bold on
                    the section's primary fill colour.
        Row N+1..N+3 -- the big value, merged across three rows for
                    visual mass. 28pt bold, centered.
        Row N+4  -- descriptor (e.g. "Reporting Population"), small
                    grey text on a soft variant of the accent colour.

    Cells are merged across `col_left:col_right` so the card spans
    the full assigned column range regardless of column width.
    """
    primary_fill = PatternFill("solid", fgColor=accent_primary)
    soft_fill = PatternFill("solid", fgColor=accent_soft)
    centered = Alignment(horizontal="center", vertical="center")

    # Row N+0: section label band.
    label_range = f"{col_left}{top_row}:{col_right}{top_row}"
    ws.merge_cells(label_range)
    lab = ws[f"{col_left}{top_row}"]
    lab.value = section_label.upper()
    lab.font = Font(bold=True, color="FFFFFF", size=11)
    lab.alignment = centered
    lab.fill = primary_fill
    ws.row_dimensions[top_row].height = 22

    # Rows N+1..N+3: value (merged 3 rows tall for visual weight).
    value_range = f"{col_left}{top_row + 1}:{col_right}{top_row + 3}"
    ws.merge_cells(value_range)
    v = ws[f"{col_left}{top_row + 1}"]
    v.value = value
    v.number_format = number_format
    v.font = Font(bold=True, size=28, color="1F4E79")
    v.alignment = centered
    v.fill = soft_fill
    for r in (top_row + 1, top_row + 2, top_row + 3):
        ws.row_dimensions[r].height = 20

    # Row N+4: descriptor.
    desc_range = f"{col_left}{top_row + 4}:{col_right}{top_row + 4}"
    ws.merge_cells(desc_range)
    d = ws[f"{col_left}{top_row + 4}"]
    d.value = descriptor
    d.font = Font(italic=True, color="555555", size=10)
    d.alignment = centered
    d.fill = soft_fill
    ws.row_dimensions[top_row + 4].height = 18


def _write_kpi_section(ws, start_row, col_left, col_right,
                        title, accent_primary, accent_soft, metrics):
    """Render a section header band + a vertical list of (label,
    value, number_format) tuples.

    Lays out one of the four detail panels (Workforce / Attendance
    Risk / Payroll Impact / Data Quality) as a small 2-column table:
    the LEFT half of the merged cell range carries the metric label,
    the RIGHT half carries the value. Zebra striping is applied to
    every other body row for readability.

    Returns the row right after the last written row (1 blank gap).
    """
    from openpyxl.utils import column_index_from_string, get_column_letter
    centered = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    band = PatternFill("solid", fgColor="F8F8F8")
    primary_fill = PatternFill("solid", fgColor=accent_primary)
    soft_fill = PatternFill("solid", fgColor=accent_soft)

    # Title band (full width of the panel).
    title_range = f"{col_left}{start_row}:{col_right}{start_row}"
    ws.merge_cells(title_range)
    th = ws[f"{col_left}{start_row}"]
    th.value = title
    th.font = Font(bold=True, color="FFFFFF", size=12)
    th.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    th.fill = primary_fill
    ws.row_dimensions[start_row].height = 22

    # Body rows: split panel into label + value columns at midpoint.
    left_idx = column_index_from_string(col_left)
    right_idx = column_index_from_string(col_right)
    mid_idx = (left_idx + right_idx) // 2
    label_left, label_right = col_left, get_column_letter(mid_idx)
    value_left, value_right = get_column_letter(mid_idx + 1), col_right

    for i, (label, value, fmt) in enumerate(metrics):
        r = start_row + 1 + i
        # Label half.
        ws.merge_cells(f"{label_left}{r}:{label_right}{r}")
        lc = ws[f"{label_left}{r}"]
        lc.value = label
        lc.font = Font(size=10)
        lc.alignment = left_align
        # Value half.
        ws.merge_cells(f"{value_left}{r}:{value_right}{r}")
        vc = ws[f"{value_left}{r}"]
        vc.value = value
        vc.number_format = fmt
        vc.font = Font(bold=True, size=10)
        vc.alignment = centered
        if i % 2 == 1:
            lc.fill = band
            vc.fill = band
        ws.row_dimensions[r].height = 18

    # Light coloured underline row to visually close the panel.
    end_row = start_row + 1 + len(metrics)
    ws.merge_cells(f"{col_left}{end_row}:{col_right}{end_row}")
    end_cell = ws[f"{col_left}{end_row}"]
    end_cell.fill = soft_fill
    ws.row_dimensions[end_row].height = 4
    return end_row + 1   # +1 for visual gap


def _write_dashboard_backing_tables(wb, summary):
    """Write the 3 backing tables to a HIDDEN helper sheet.

    Why: the pie/bar charts on the Dashboard reference these rows via
    openpyxl `Reference` objects. Keeping them on the Dashboard itself
    forced an "Underlying Data" zone at the bottom of the visible
    sheet, cluttering the executive view. The polish pass moves them
    onto a dedicated `Dashboard Data` sheet that is created on demand,
    marked hidden, and never seen by HR/payroll users -- it shows up
    in Excel's "Unhide sheet" dialog if anyone genuinely needs to
    audit the chart inputs.

    Returns a dict that includes the helper worksheet plus the row
    ranges for each table so `_write_dashboard_charts` can build
    cross-sheet `Reference` objects:

        {
          "ws": <helper Worksheet>,
          "status":      (data_start, data_end, n_cols),
          "overtime":    (data_start, data_end, simplified_df),
          "early_leave": (data_start, data_end, simplified_df),
        }
    """
    if _DASH_HELPER_SHEET_NAME in wb.sheetnames:
        helper = wb[_DASH_HELPER_SHEET_NAME]
    else:
        helper = wb.create_sheet(_DASH_HELPER_SHEET_NAME)
    helper.sheet_state = "hidden"
    helper.sheet_view.showGridLines = False

    cursor = 1
    helper.cell(row=cursor, column=1,
                value="Underlying Data (used by Dashboard charts)").font = (
        _SECTION_FONT
    )
    cursor += 2

    # Attendance Status (drives the pie chart).
    helper.cell(row=cursor, column=1,
                value="Attendance Status").font = _SECTION_FONT
    status_df = summary["status_summary"]
    _, status_start, status_end, status_next, status_n_cols = _write_dataframe(
        helper, status_df, cursor + 1,
    )
    _format_numeric_cells(
        helper,
        rows=range(status_start, status_end + 1),
        cols=range(2, status_n_cols + 1),
        number_format="#,##0",
    )

    # Top Overtime (drives the bar chart).
    # The chart's title/subtitle/X-axis title all promise PAYABLE
    # OVERTIME HOURS at the 1.5x multiplier. The backing table must
    # therefore expose `total_overtime_payable_hours` so the chart
    # Reference can point at the right unit. Earlier versions of this
    # function only carried minutes + actual-hours and the chart was
    # silently plotting minutes under a "Hours" title -- off by 60x.
    overtime_cols = [
        "Employee ID", "First Name",
        "total_overtime_minutes", "total_overtime_hours",
        "total_overtime_payable_minutes", "total_overtime_payable_hours",
    ]
    top_overtime_full = summary.get("top_overtime_employees")
    if top_overtime_full is not None and not top_overtime_full.empty:
        # Fallback path: when the upstream aggregate doesn't have the
        # payable columns (e.g. multiplier disabled in a non-default
        # config) fill them with NaN so the chart renders cleanly
        # without crashing AND without lying about the unit.
        simplified_overtime = top_overtime_full.head(10).copy()
        for col in overtime_cols:
            if col not in simplified_overtime.columns:
                simplified_overtime[col] = pd.NA
        simplified_overtime = simplified_overtime[overtime_cols]
    else:
        simplified_overtime = pd.DataFrame(columns=overtime_cols)
    helper.cell(row=status_next, column=1,
                value="Top Overtime Employees").font = _SECTION_FONT
    _, ot_start, ot_end, ot_next, _ = _write_dataframe(
        helper, simplified_overtime, status_next + 1,
    )
    if not simplified_overtime.empty:
        # Integer minutes columns (3, 5) -- thousand separators.
        _format_numeric_cells(
            helper, rows=range(ot_start, ot_end + 1),
            cols=[3, 5], number_format="#,##0",
        )
        # Hour columns (4, 6) -- 1 decimal place. The chart Reference
        # below points at column 6 (`total_overtime_payable_hours`) so
        # the bar data labels also render with 1 decimal precision.
        _format_numeric_cells(
            helper, rows=range(ot_start, ot_end + 1),
            cols=[4, 6], number_format="0.0",
        )

    # Top Early Leave (drives the bar chart).
    top_el_full = summary.get("top_early_leave_employees")
    el_cols = [
        "Employee ID", "First Name",
        "total_early_leave_minutes", "total_early_leave_hours",
    ]
    if top_el_full is not None and not top_el_full.empty:
        simplified_el = top_el_full[el_cols].head(10).copy()
    else:
        simplified_el = pd.DataFrame(columns=el_cols)
    helper.cell(row=ot_next, column=1,
                value="Top Early Leave Employees").font = _SECTION_FONT
    _, el_start, el_end, _, _ = _write_dataframe(
        helper, simplified_el, ot_next + 1,
    )
    if not simplified_el.empty:
        _format_numeric_cells(
            helper, rows=range(el_start, el_end + 1),
            cols=[3], number_format="#,##0",
        )
        _format_numeric_cells(
            helper, rows=range(el_start, el_end + 1),
            cols=[4], number_format="0.0",
        )

    return {
        "ws": helper,
        "status":      (status_start, status_end, status_n_cols),
        "overtime":    (ot_start, ot_end, simplified_overtime),
        "early_leave": (el_start, el_end, simplified_el),
    }


def _write_dashboard_charts(ws, wb, backing, other_sheet_first_data_row):
    """Add the 5 chart objects.

    Anchors align with the new card grid (cards in A:L within A:M):
        anchor_tl ("A{row}")   above cards 1-2     (Attendance Status pie)
        anchor_tr ("H{row}")   above cards 3-4     (Daily Late Trend line)
        anchor_ml ("A{row+s}") second-row left     (Top Late bar)
        anchor_mr ("H{row+s}") second-row right    (Top Overtime bar)
        anchor_bl ("A{row+2s}") third-row left     (Top Early Leave bar)
    Charts are 16cm x 9cm so each row of charts is ~22 rows tall;
    `_DASH_CHART_ROW_STRIDE` keeps them spaced cleanly.

    `backing` is the dict returned by
    `_write_dashboard_backing_tables`; its `ws` is the HIDDEN helper
    sheet, so on-sheet refs point there (not at the visible Dashboard).
    """
    base = _DASH_CHARTS_ROW
    stride = _DASH_CHART_ROW_STRIDE
    anchor_tl = f"A{base}"
    anchor_tr = f"H{base}"
    anchor_ml = f"A{base + stride}"
    anchor_mr = f"H{base + stride}"
    anchor_bl = f"A{base + 2 * stride}"

    helper_ws = backing["ws"]
    status_start, status_end, _ = backing["status"]
    ot_start, ot_end, simplified_overtime = backing["overtime"]
    el_start, el_end, simplified_el = backing["early_leave"]

    # Chart 1 (top-left): Attendance Status pie.
    pie_labels = Reference(helper_ws, min_col=1,
                           min_row=status_start, max_row=status_end)
    pie_values = Reference(helper_ws, min_col=2,
                           min_row=status_start, max_row=status_end)
    ws.add_chart(
        _pie_chart(
            "Attendance Status Breakdown",
            pie_labels, pie_values,
            subtitle="(Distribution %)",
        ),
        anchor_tl,
    )

    # Chart 2 (top-right): Daily Late Trend (references Daily Trend sheet).
    if "Daily Trend" in wb.sheetnames:
        trend_ws = wb["Daily Trend"]
        trend_first_data = other_sheet_first_data_row
        trend_header_row = trend_first_data - 1
        if trend_ws.max_row >= trend_first_data:
            trend_labels = Reference(
                trend_ws, min_col=1,
                min_row=trend_first_data, max_row=trend_ws.max_row,
            )
            trend_values = Reference(
                trend_ws, min_col=3,
                min_row=trend_header_row, max_row=trend_ws.max_row,
            )
            ws.add_chart(
                _line_chart(
                    "Daily Late Trend",
                    trend_labels, trend_values,
                    subtitle="(Number of Late Cases)",
                    x_axis_title="Date",
                    y_axis_title="Late Cases",
                ),
                anchor_tr,
            )

    # Chart 3 (middle-left): Top Late Employees -- references Employee Summary.
    emp_ws = wb["Employee Summary"]
    emp_first_data = other_sheet_first_data_row
    if emp_ws.max_row >= emp_first_data:
        n_late = min(10, emp_ws.max_row - emp_first_data + 1)
        late_labels = Reference(
            emp_ws, min_col=2,
            min_row=emp_first_data, max_row=emp_first_data + n_late - 1,
        )
        late_values = Reference(
            emp_ws, min_col=7,
            min_row=emp_first_data, max_row=emp_first_data + n_late - 1,
        )
        ws.add_chart(
            _bar_chart(
                "Top Late Employees", late_labels, late_values,
                horizontal=True,
                subtitle="By Total Late Hours (Hours)",
                x_axis_title="Late Hours (Hours)",
            ),
            anchor_ml,
        )

    # Chart 4 (middle-right): Top Overtime Employees.
    # IMPORTANT: the value Reference MUST point at column 6
    # (`total_overtime_payable_hours`), not column 3 (minutes). The
    # chart's title, subtitle, and X-axis title all promise PAYABLE
    # HOURS at the 1.5x multiplier; plotting minutes under that label
    # was a 60x unit error caught in the accuracy audit.
    if ot_end >= ot_start and not simplified_overtime.empty:
        ot_labels = Reference(helper_ws, min_col=2,
                              min_row=ot_start, max_row=ot_end)
        ot_values = Reference(helper_ws, min_col=6,
                              min_row=ot_start, max_row=ot_end)
        ws.add_chart(
            _bar_chart(
                "Top Overtime Employees", ot_labels, ot_values,
                horizontal=True,
                subtitle="By Total Overtime Hours (Payable 1.5x)",
                x_axis_title="Overtime Hours (Payable 1.5x)",
            ),
            anchor_mr,
        )

    # Chart 5 (bottom-left): Top Early Leave Employees.
    if el_end >= el_start and not simplified_el.empty:
        el_labels = Reference(helper_ws, min_col=2,
                              min_row=el_start, max_row=el_end)
        el_values = Reference(helper_ws, min_col=3,
                              min_row=el_start, max_row=el_end)
        ws.add_chart(
            _bar_chart(
                "Top Early Leave Employees", el_labels, el_values,
                horizontal=True,
                subtitle="By Total Early Leave Minutes",
                x_axis_title="Early Leave Minutes",
            ),
            anchor_bl,
        )

    # ---- "How to read" annotations for the 3 bar charts --------------
    # Positioned just below each chart row, inside the gap between the
    # chart's bottom edge and the next chart-row anchor. Accent colour
    # ties each annotation back to its KPI section (Attendance Risk
    # red for late metrics, Payroll Impact green for overtime).
    risk_primary, _ = _DASH_SECTION_COLORS["Attendance Risk"]
    payroll_primary, _ = _DASH_SECTION_COLORS["Payroll Impact"]
    # Charts in chart-row 2 (anchored at base+stride) extend ~18 rows
    # downward; place annotation 2 rows below the anchor + chart span.
    annotation_row_2 = base + stride + 18
    annotation_row_3 = base + 2 * stride + 18
    _write_chart_annotation(
        ws, top_row=annotation_row_2,
        col_left="A", col_right="F",
        chart_label="Top Late Employees",
        body=(
            "Shows the 10 employees with the highest total late hours "
            "during the reporting period. Higher is worse -- these "
            "names drive the Attendance Risk KPI on the cards above."
        ),
        accent_primary=risk_primary,
    )
    _write_chart_annotation(
        ws, top_row=annotation_row_2,
        col_left="H", col_right="M",
        chart_label="Top Overtime Employees",
        body=(
            "Shows the 10 employees with the highest payable overtime "
            "hours (1.5x multiplier applied). Higher means higher "
            "payroll impact -- these drive the Payroll Impact KPI."
        ),
        accent_primary=payroll_primary,
    )
    _write_chart_annotation(
        ws, top_row=annotation_row_3,
        col_left="A", col_right="F",
        chart_label="Top Early Leave Employees",
        body=(
            "Shows the 10 employees with the highest total early-leave "
            "minutes during the reporting period. Excused minutes have "
            "already been subtracted (only unexcused early leave shown)."
        ),
        accent_primary=risk_primary,
    )


def _build_dashboard(wb, summary, period_start=None, period_end=None):
    """Compose the Dashboard sheet from small, focused helpers.

    Layout (executive-friendly, top-to-bottom):
      1. Title + Reporting Period banner   (rows 1-5)
      2. 4 hero KPI cards                  (rows 6-10)
      3. Charts (5-up grid)                (rows 12-68)
      4. Detail KPI sections (A/B/C/D)     (rows 70-104)
      5. Underlying Data backing tables    (rows 110+)
    """
    ws = wb["Dashboard"]
    # Clean executive look: no gridlines anywhere on the dashboard.
    ws.sheet_view.showGridLines = False

    # ----- Zone 1: Title + Reporting Period banner --------------------
    cards_row = _write_dashboard_title(ws, period_start, period_end)

    # ----- Zone 2: 4 hero KPI cards -----------------------------------
    # Pick the single MOST important metric per section. Detail metrics
    # within each section are surfaced in Zone 4 below.
    hero_metrics = [
        ("Workforce",       "Reporting Population",
         summary.get("reporting_population",
                     summary.get("total_employees", 0)),
         "#,##0"),
        ("Attendance Risk", "Late Cases",
         summary.get("late_cases", 0), "#,##0"),
        ("Payroll Impact",  "Total Over Time (Payable 1.5x) (Hours)",
         summary.get("total_overtime_payable_hours", 0), "0.00"),
        ("Data Quality",    "Data Quality Score",
         summary.get("data_quality_score", 0), "0.0"),
    ]
    for (col_left, col_right), (section, descriptor, value, fmt) in zip(
        _DASH_CARD_COLS, hero_metrics,
    ):
        primary, soft = _DASH_SECTION_COLORS[section]
        _write_kpi_card(
            ws, top_row=cards_row,
            col_left=col_left, col_right=col_right,
            section_label=section, descriptor=descriptor,
            value=value, number_format=fmt,
            accent_primary=primary, accent_soft=soft,
        )

    # ----- Zone 3: Charts -- written AFTER backing tables exist -------
    # Backing tables must exist BEFORE chart objects can reference
    # their cells. They are written to a HIDDEN helper sheet
    # ("Dashboard Data") so they no longer occupy visible rows on
    # the Dashboard itself.
    backing = _write_dashboard_backing_tables(wb, summary)
    show_period = _has_period(period_start, period_end)
    other_sheet_first_data_row = _PERIOD_BLOCK_ROWS + 2 if show_period else 2
    _write_dashboard_charts(ws, wb, backing, other_sheet_first_data_row)

    # ----- Zone 4: Detail KPI sections (A/B/C/D) ----------------------
    sections = [
        ("A. Workforce", "Workforce", [
            ("Reporting Population",
             summary.get("reporting_population",
                         summary.get("total_employees", 0)),
             "#,##0"),
            ("Excluded Employees",
             summary.get("excluded_employee_count", 0), "#,##0"),
            ("Aliases Used",
             summary.get("employee_id_aliases_used", 0), "#,##0"),
        ]),
        ("B. Attendance Risk", "Attendance Risk", [
            ("Late Cases", summary.get("late_cases", 0), "#,##0"),
            ("Total Late Minutes (Unexcused)",
             summary.get("total_late_minutes", 0), "#,##0"),
            ("Approved Excuse Cases",
             summary.get("approved_excuse_cases", 0), "#,##0"),
            ("Leave Cases", summary.get("leave_cases", 0), "#,##0"),
            ("High Risk Employees",
             summary.get("high_risk_employees", 0), "#,##0"),
            ("Missing Check-Out Cases",
             summary.get("missing_check_out_cases", 0), "#,##0"),
            ("Early Leave Cases",
             summary.get("early_leave_cases", 0), "#,##0"),
            ("Total Early Leave Minutes",
             summary.get("total_early_leave_minutes", 0), "#,##0"),
            ("Early Leave Anomalies (review)",
             summary.get("early_leave_anomaly_cases", 0), "#,##0"),
        ]),
        ("C. Payroll Impact", "Payroll Impact", [
            ("Overtime Cases", summary.get("overtime_cases", 0), "#,##0"),
            ("Total Over Time (Hours) (Actual)",
             summary.get("total_overtime_hours", 0), "0.0"),
            ("Overtime Multiplier",
             summary.get("overtime_multiplier", 1.0), "0.00"),
            ("Total Over Time (Payable 1.5x) (Hours)",
             summary.get("total_overtime_payable_hours", 0), "0.00"),
            ("Estimated Deduction (capped)",
             summary.get("total_deduction_capped", 0), "#,##0.00"),
        ]),
        ("D. Data Quality / Notes", "Data Quality", [
            ("Data Quality Score",
             summary.get("data_quality_score", 0), "0.0"),
            ("Missing Schedule Cases",
             summary.get("missing_schedule_cases", 0), "#,##0"),
            ("Total Break Count (info)",
             summary.get("total_break_count", 0), "#,##0"),
            ("Total Break Minutes (info)",
             summary.get("total_break_minutes", 0), "#,##0"),
            ("Employees With Breaks (info)",
             summary.get("employees_with_breaks", 0), "#,##0"),
            ("Incomplete Break Records (info)",
             summary.get("incomplete_break_records", 0), "#,##0"),
            ("Alias Records Mapped (info)",
             summary.get("employee_id_alias_records_mapped", 0), "#,##0"),
        ]),
    ]
    # Lay out the 4 sections in a 2x2 grid: A and B side-by-side, then
    # C and D side-by-side below them. The left/right vertical cursors
    # advance independently so sections of unequal length never overlap.
    left_cursor = right_cursor = _DASH_DETAIL_SECTIONS_ROW
    for idx, (title, section_key, metrics) in enumerate(sections):
        primary, soft = _DASH_SECTION_COLORS[section_key]
        col_left, col_right = _DASH_DETAIL_COLS[idx % 2]
        row = left_cursor if idx % 2 == 0 else right_cursor
        next_row = _write_kpi_section(
            ws, start_row=row,
            col_left=col_left, col_right=col_right,
            title=title,
            accent_primary=primary, accent_soft=soft,
            metrics=metrics,
        )
        if idx % 2 == 0:
            left_cursor = next_row
        else:
            right_cursor = next_row

    # ----- Column widths + freeze -------------------------------------
    # All visible content lives within A:M. Uniform 14-wide columns
    # give the 4 cards equal weight and let the title / banner / detail
    # panels merge cleanly across the full visible width.
    for letter in ("A", "B", "C", "D", "E", "F", "G", "H",
                   "I", "J", "K", "L", "M"):
        ws.column_dimensions[letter].width = 14
    # Freeze just under the cards so title + banner + cards stay
    # visible when the reader scrolls down to charts or details.
    ws.freeze_panes = f"A{cards_row + _DASH_CARD_HEIGHT}"


def export_report(
    summary, daily, raw_punches=None, schedules_df=None,
    period_start=None, period_end=None,
):
    """Write the monthly Excel workbook.

    `raw_punches` is the post-alias, post-manual-correction punch
    events dataframe (one row per Check In / Check Out / Break event).
    When provided alongside `schedules_df`, the Employee Attendance
    audit sheet is rendered with paired-shift columns. Both arguments
    are optional for backward compat -- old callers that pass only
    `summary` + `daily` keep working and simply skip the new sheet.

    `period_start` / `period_end` drive the Reporting Period banner at
    the top of every executive sheet (Dashboard, Employee Summary,
    Employee Attendance, Daily Attendance, Daily Trend, Department
    Summary, and every audit sheet). When both are None the banner is
    suppressed and sheets keep the legacy row-1 header.
    """
    print("Exporting Excel report...")

    wb = Workbook()
    wb.active.title = "Dashboard"

    # Pre-resolve the period-banner kwargs once so every sheet builder
    # call below stays uniform and the banner can be turned off by a
    # single `None` pair in tests / legacy callers.
    pkw = {"period_start": period_start, "period_end": period_end}
    daily_data_start = (
        _PERIOD_BLOCK_ROWS + 2 if _has_period(period_start, period_end) else 2
    )

    # Always-present data sheets, in this tab order.
    _build_executive_employee_sheet(
        wb.create_sheet("Employee Summary"),
        summary.get("executive_employee_summary"),
        **pkw,
    )
    # Employee Attendance -- audit-focused daily summary with split-
    # shift columns. Placed right after Employee Summary so HR can
    # cross-reference the executive totals with the per-day source.
    if raw_punches is not None and schedules_df is not None:
        attendance_rows = build_employee_attendance_rows(
            raw_punches, schedules_df,
        )
        _build_employee_attendance_sheet(
            wb.create_sheet("Employee Attendance"),
            attendance_rows,
            **pkw,
        )
    daily_ws = wb.create_sheet("Daily Attendance")
    _build_data_sheet(daily_ws, daily, **pkw)
    _apply_daily_conditional_formatting(
        daily_ws, daily, data_start_row=daily_data_start,
    )
    _build_data_sheet(
        wb.create_sheet("Daily Trend"), summary.get("daily_trend"), **pkw,
    )

    # Optional sheets -- only added when the source data supplied them.
    missing_punches = summary.get("missing_punch_summary")
    if missing_punches is not None and not missing_punches.empty:
        _build_data_sheet(
            wb.create_sheet("Missing Punches"), missing_punches, **pkw,
        )

    department_summary = summary.get("department_summary")
    if department_summary is not None and not department_summary.empty:
        _build_data_sheet(
            wb.create_sheet("Department Summary"), department_summary, **pkw,
        )

    reconciliation_details = summary.get("employee_reconciliation_details")
    if reconciliation_details is not None and not reconciliation_details.empty:
        _build_data_sheet(
            wb.create_sheet("Employee Reconciliation Details"),
            reconciliation_details, **pkw,
        )

    employee_master = summary.get("employee_master")
    if employee_master is not None and not employee_master.empty:
        _build_data_sheet(
            wb.create_sheet("Employee Master"), employee_master, **pkw,
        )

    # Overtime sheet -- only rows where overtime actually happened.
    overtime_rows = daily[daily.get("overtime_status") == "Overtime"]
    if not overtime_rows.empty:
        _build_data_sheet(
            wb.create_sheet("Overtime"),
            overtime_rows[[
                "Employee ID", "First Name", "Date",
                "Check In", "Check Out", "Shift Start", "Shift End",
                "matched_shift_start", "matched_shift_end",
                "matched_shift_label", "shift_intervals",
                "worked_minutes", "scheduled_minutes",
                "matched_scheduled_minutes",
                "overtime_minutes", "overtime_status",
                "overtime_policy", "overtime_calculation_note",
                "overtime_multiplier",
                "overtime_payable_minutes", "overtime_payable_hours",
            ]],
            **pkw,
        )

    # Early Leave sheet -- only rows where the employee genuinely left early.
    early_leave_rows = daily[daily.get("early_leave_status") == "Early Leave"]
    if not early_leave_rows.empty:
        el_cols = [
            "Employee ID", "First Name", "Date",
            "Check In", "Check Out", "Shift Start", "Shift End",
            "matched_shift_start", "matched_shift_end",
            "matched_shift_label", "shift_intervals",
            "matched_scheduled_minutes",
            "early_leave_minutes", "early_leave_status",
            "early_leave_anomaly", "early_leave_anomaly_reason",
        ]
        # Excused / unexcused early-leave columns are appended when
        # _attach_excused_early_leave_info ran (which it does in the
        # standard pipeline). HR uses these to see which portion of
        # the gap was covered by an approved permission.
        for extra in ("excused_early_leave_minutes",
                      "unexcused_early_leave_minutes"):
            if extra in early_leave_rows.columns:
                el_cols.append(extra)
        _build_data_sheet(
            wb.create_sheet("Early Leave"),
            early_leave_rows[el_cols],
            **pkw,
        )

    excluded_summary = summary.get("excluded_employees_summary")
    if excluded_summary is not None and not excluded_summary.empty:
        _build_data_sheet(
            wb.create_sheet("Excluded Employees"), excluded_summary, **pkw,
        )

    # Break analytics sheet -- informational, only when breaks exist.
    break_summary = summary.get("break_summary")
    if break_summary is not None and not break_summary.empty:
        _build_data_sheet(
            wb.create_sheet("Break Summary"), break_summary, **pkw,
        )

    # Absence audit ledger -- every (employee, date) and why it was
    # or wasn't counted as an absence.
    absence_details = summary.get("absence_details")
    if absence_details is not None and not absence_details.empty:
        _build_data_sheet(
            wb.create_sheet("Absence Details"), absence_details, **pkw,
        )

    # Per-employee audit totals (scheduled / attended / permission /
    # vacation / secondment / absence) with a reconciliation_delta so
    # HR can spot bookkeeping inconsistencies at a glance.
    absence_audit = summary.get("absence_audit")
    if absence_audit is not None and not absence_audit.empty:
        _build_data_sheet(
            wb.create_sheet("Absence Audit"), absence_audit, **pkw,
        )

    # Employee ID alias audit -- which historical IDs got remapped to
    # current IDs (only when at least one alias was configured active).
    alias_audit = summary.get("alias_audit")
    if alias_audit is not None and not alias_audit.empty:
        _build_data_sheet(
            wb.create_sheet("Employee ID Alias Audit"), alias_audit, **pkw,
        )

    # Schedule lookup audit -- one row per (Employee ID, attendance name)
    # describing how the Odoo schedule was matched (or why not). Always
    # emitted so HR can audit Missing Schedule rows even when zero rows
    # are missing.
    schedule_audit = summary.get("schedule_lookup_audit")
    if schedule_audit is not None and not schedule_audit.empty:
        _build_data_sheet(
            wb.create_sheet("Schedule Lookup Audit"), schedule_audit, **pkw,
        )

    # Manual punch corrections that did NOT clear the approval / evidence
    # gates -- exposed for Exceptions & Manual Review follow-up.
    rejected_corrections = summary.get("rejected_punch_corrections")
    if rejected_corrections is not None and not rejected_corrections.empty:
        _build_data_sheet(
            wb.create_sheet("Manual Punch Rejections"),
            rejected_corrections, **pkw,
        )

    # High-level reconciliation table lives on its own sheet so the
    # Dashboard stays uncluttered.
    reconciliation = summary.get("employee_reconciliation")
    if reconciliation is not None and not reconciliation.empty:
        _build_data_sheet(
            wb.create_sheet("Reconciliation"), reconciliation, **pkw,
        )

    # Build Dashboard LAST so it can reference positions on the other sheets.
    _build_dashboard(wb, summary, **pkw)

    now = datetime.now()
    monthly_dir = REPORT_OUTPUT_DIR / now.strftime("%Y-%m")
    monthly_dir.mkdir(parents=True, exist_ok=True)
    filename = monthly_dir / f"hr_report_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)

    print(f"Report saved: {filename}")
