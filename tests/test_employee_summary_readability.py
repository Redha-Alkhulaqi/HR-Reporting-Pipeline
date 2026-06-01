"""Phase-1 readability refresh for the Employee Summary sheet.

These tests pin the layout decisions that improve readability for
HR/payroll users WITHOUT changing any calculation or column. They are
intentionally orthogonal to:

  - tests/test_executive_summary.py    (the 14-column schema + values)
  - tests/test_employee_id_text_format.py (Employee ID TEXT format)
  - tests/test_reporting_period_header.py (banner contract)

so a future tweak to the readability layer (e.g. swapping a tint
shade) doesn't accidentally regress the column count, business
totals, or banner contract.
"""
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from excel_exporter import (
    _EMP_SUMMARY_IMPORTANT_TINTS,
    _EMP_SUMMARY_ZEBRA_FILL,
    _build_executive_employee_sheet,
    export_report,
)


PERIOD_START = "2026-04-20"
PERIOD_END = "2026-05-19"


def _executive_df(n_rows=4):
    """Build a small Employee Summary frame with `n_rows` employees."""
    rows = []
    for i in range(n_rows):
        rows.append({
            "Employee ID": 4195162 + i,
            "First Name": f"EMP{i + 1}",
            "No of Absence Days": 1.0 * i,
            "No of Permission Days": i,
            "No of Vacation Days": 0,
            "No of Secondment Days": 0,
            "Total Late (Hours)": 0.5 * i,
            "Total Over Time (Hours) (Actual)": 2.0,
            "Total Over Time (Payable 1.5x) (Hours)": 3.0,
            "Total Early Leave (Hours)": 0.1 * i,
            "Break Time (Hours)": 0.0,
            "Break Time (After Policy)": 0.0,
            "Friday Compensation Days": 0,
            "Friday Worked Dates": "",
        })
    return pd.DataFrame(rows)


# ---- Header band ---------------------------------------------------------

def test_header_row_uses_wrap_text_and_taller_height():
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(
        ws, _executive_df(2),
        period_start=PERIOD_START, period_end=PERIOD_END,
    )
    # Header lands at row 4 (with the period banner above).
    header_row = 4
    for c in range(1, 15):
        cell = ws.cell(row=header_row, column=c)
        assert cell.alignment.wrap_text is True, (
            f"column {c} header missing wrap_text"
        )
    # Header row height bumped so the wrapped 2-3 line labels render.
    assert ws.row_dimensions[header_row].height >= 36


# ---- Freeze pane: header row + first two columns -------------------------

def test_freeze_pane_covers_header_row_and_first_two_columns_with_banner():
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(
        ws, _executive_df(2),
        period_start=PERIOD_START, period_end=PERIOD_END,
    )
    # With banner: header at row 4 -> freeze at "C5" (rows 1-4 + cols A,B).
    assert ws.freeze_panes == "C5"


def test_freeze_pane_covers_first_two_columns_without_banner():
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(ws, _executive_df(2))
    # No banner: header at row 1 -> freeze at "C2" (row 1 + cols A,B).
    assert ws.freeze_panes == "C2"


# ---- Numeric formats consistent across data rows -------------------------

def test_integer_day_count_columns_use_thousands_format():
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(
        ws, _executive_df(3),
        period_start=PERIOD_START, period_end=PERIOD_END,
    )
    # Integer day-count cols: 4 (Permission), 5 (Vacation),
    # 6 (Secondment), 13 (Friday Compensation Days).
    for row in (5, 6, 7):           # data rows when banner is on
        for col in (4, 5, 6, 13):
            assert ws.cell(row=row, column=col).number_format == "#,##0"


def test_hour_and_fractional_columns_use_one_decimal_format():
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(
        ws, _executive_df(3),
        period_start=PERIOD_START, period_end=PERIOD_END,
    )
    for row in (5, 6, 7):
        for col in (3, 7, 8, 9, 10, 11, 12):
            assert ws.cell(row=row, column=col).number_format == "0.0"


# ---- Important-column tints + zebra striping -----------------------------

def _fg(cell):
    """Return the foreground hex of a cell's fill, normalized to uppercase
    and 6 chars (drops the alpha prefix openpyxl adds for theme fills)."""
    fg = cell.fill.fgColor
    val = fg.value or ""
    if isinstance(val, str) and len(val) == 8:
        val = val[2:]
    return val.upper()


def _is_no_fill(cell):
    """openpyxl returns one of '', '000000', or '00000000' when no
    fill has been set explicitly. Normalize to a single check."""
    return _fg(cell) in ("", "000000")


@pytest.mark.parametrize("col, expected_hex", [
    (3,  "FCE4E4"),   # Absence
    (7,  "FBE5D6"),   # Late Hours
    (10, "FFF2CC"),   # Early Leave
    (12, "DDEBF7"),   # Break After Policy
])
def test_important_columns_tinted_on_every_data_row(col, expected_hex):
    """The 4 "important" payroll columns must carry the same soft tint
    on EVERY data row so HR's eye lands on them column-wise. Zebra
    striping for the rest of the columns must NOT override these tints."""
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(
        ws, _executive_df(4),
        period_start=PERIOD_START, period_end=PERIOD_END,
    )
    # Data rows: 5, 6, 7, 8 (banner pushes header to row 4).
    for r in (5, 6, 7, 8):
        assert _fg(ws.cell(row=r, column=col)) == expected_hex, (
            f"col {col} row {r} tint mismatch"
        )


def test_payable_overtime_column_has_its_own_bolder_green_tint():
    """Column 9 (Payable OT) keeps its prominent green + bold + border
    treatment from _apply_payable_overtime_styling and is NOT overwritten
    by the zebra/important-column pass."""
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(
        ws, _executive_df(3),
        period_start=PERIOD_START, period_end=PERIOD_END,
    )
    for r in (5, 6, 7):
        cell = ws.cell(row=r, column=9)
        # The light-green body fill from _apply_payable_overtime_styling.
        assert _fg(cell) == "E2EFDA"
        assert cell.font.bold is True


def test_non_highlighted_columns_get_zebra_stripe_on_even_data_rows():
    """Non-important columns alternate between no fill (odd rows) and
    the subtle gray zebra fill (even rows)."""
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(
        ws, _executive_df(4),
        period_start=PERIOD_START, period_end=PERIOD_END,
    )
    # Column 2 (First Name) is not in the important set and not col 9.
    # Banner-on data rows: 5 (odd), 6 (even), 7 (odd), 8 (even).
    assert _is_no_fill(ws.cell(row=5, column=2))
    assert _fg(ws.cell(row=6, column=2)) == _EMP_SUMMARY_ZEBRA_FILL
    assert _is_no_fill(ws.cell(row=7, column=2))
    assert _fg(ws.cell(row=8, column=2)) == _EMP_SUMMARY_ZEBRA_FILL


def test_important_column_set_matches_user_spec():
    """Spec-of-record: the readability refresh highlights exactly these
    4 columns + the Payable OT hero column. Locking the constant so
    future tweaks have to update this test deliberately."""
    assert set(_EMP_SUMMARY_IMPORTANT_TINTS.keys()) == {3, 7, 10, 12}


# ---- Compact Notes block -------------------------------------------------

def test_notes_block_is_compact_two_rows_max():
    """The pre-Phase-1 Notes block was 9 rows of bullets + a badge,
    which dominated the bottom of the sheet on a laptop screen. The
    refresh collapses it to a Notes title + a single wrapped body
    sentence (2 visible rows, plus the surrounding spacer)."""
    wb = Workbook()
    ws = wb.active
    df = _executive_df(2)
    _build_executive_employee_sheet(
        ws, df, period_start=PERIOD_START, period_end=PERIOD_END,
    )
    # Banner=3 + header=1 + data=2 = max_row 6 before notes.
    # Notes block adds 1 spacer + 2 rows = max_row 9.
    # Locate the Notes title.
    notes_title_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v == "Notes":
            notes_title_row = r
            break
    assert notes_title_row is not None, "Notes title row missing"
    # Body sits one row below the title.
    body = ws.cell(row=notes_title_row + 1, column=1).value
    assert isinstance(body, str)
    assert "Payable Overtime" in body
    assert "1.5" in body
    # Nothing further -- no extra bullet rows.
    extra = ws.cell(row=notes_title_row + 2, column=1).value
    assert extra in (None, ""), (
        f"unexpected extra Notes content at row {notes_title_row + 2}: {extra!r}"
    )


# ---- Print-friendly settings ---------------------------------------------

def test_print_settings_are_landscape_fit_to_width_repeat_header():
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(
        ws, _executive_df(2),
        period_start=PERIOD_START, period_end=PERIOD_END,
    )
    assert ws.page_setup.orientation == ws.ORIENTATION_LANDSCAPE
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True
    # Repeat banner + header on every printed page (rows 1..header_row).
    # openpyxl normalizes "1:4" to "$1:$4" on save, so accept either form.
    assert ws.print_title_rows in ("1:4", "$1:$4")


def test_print_settings_without_banner_repeat_only_header():
    wb = Workbook()
    ws = wb.active
    _build_executive_employee_sheet(ws, _executive_df(2))
    assert ws.print_title_rows in ("1:1", "$1:$1")


# ---- End-to-end via export_report ----------------------------------------

def test_export_report_preserves_columns_and_print_settings(
    tmp_path, monkeypatch,
):
    """The readability refresh must not change the 14-column schema
    and must propagate the print settings through export_report."""
    monkeypatch.setattr("excel_exporter.REPORT_OUTPUT_DIR", tmp_path)
    df = _executive_df(3)
    summary = {
        "executive_employee_summary": df,
        "status_summary": pd.DataFrame(
            [{"attendance_status": "Normal", "count": 3}]
        ),
        "daily_trend": pd.DataFrame(),
        "late_cases": 0, "total_late_minutes": 0, "approved_excuse_cases": 0,
        "leave_cases": 0, "missing_schedule_cases": 0,
        "missing_check_out_cases": 0, "high_risk_employees": 0,
        "excluded_employee_count": 0, "total_deduction_capped": 0,
        "overtime_cases": 0, "total_overtime_hours": 0,
        "overtime_multiplier": 1.5, "total_overtime_payable_hours": 0,
        "early_leave_cases": 0, "total_early_leave_minutes": 0,
        "early_leave_anomaly_cases": 0, "total_break_count": 0,
        "total_break_minutes": 0, "employees_with_breaks": 0,
        "incomplete_break_records": 0, "employee_id_aliases_used": 0,
        "employee_id_alias_records_mapped": 0, "data_quality_score": 100,
        "reporting_population": 3,
    }
    daily = pd.DataFrame([{
        "Employee ID": 4195162, "First Name": "EMP1",
        "Date": "2026-05-03",
        "Check In": "08:00:00", "Check Out": "17:00:00",
        "Shift Start": "08:00:00", "Shift End": "17:00:00",
        "worked_minutes": 540, "scheduled_minutes": 540,
        "is_late": False, "unexcused_delay_minutes": 0,
        "overtime_minutes": 0, "early_leave_minutes": 0,
        "early_leave_anomaly": False, "missing_check_out": False,
        "attendance_status": "Normal", "is_excluded": False,
        "overtime_status": "Normal", "early_leave_status": "Normal",
    }])
    export_report(summary, daily,
                  period_start=PERIOD_START, period_end=PERIOD_END)

    written = list(tmp_path.rglob("hr_report_*.xlsx"))
    assert written
    wb = load_workbook(written[0])
    es = wb["Employee Summary"]
    # 14 columns still present.
    assert es.cell(row=4, column=1).value == "Employee ID"
    assert es.cell(row=4, column=14).value == "Friday Worked Dates"
    # Print settings survived.
    assert es.page_setup.orientation == "landscape"
    assert es.page_setup.fitToWidth == 1
    assert es.print_title_rows in ("1:4", "$1:$4")
    # Freeze on both axes.
    assert es.freeze_panes == "C5"
