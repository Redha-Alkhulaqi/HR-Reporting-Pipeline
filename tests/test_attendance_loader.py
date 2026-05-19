"""Attendance loader adapter tests.

BioTime can export attendance in two known templates:

1. *Transaction* (legacy):
     - 1 banner row, then headers: Employee ID, First Name, Date,
       Punch Time, Punch State.
     - One row per punch event (Check In / Check Out / Break In /
       Break Out).

2. *First In Last Out Report* (new):
     - 4 banner rows (title, period, company + generated-on, blank),
       then headers: Employee ID, First Name, Date, Weekday,
       First Check In, Last Check Out, Total Time.
     - One row per (employee, date), with pre-aggregated first
       check-in and last check-out (HH:MM, DD-MM-YYYY).

The loader must:
- Auto-detect the header row (banner length varies).
- Auto-detect the schema and convert (2) into punch-events rows
  matching schema (1) so the rest of the pipeline works unchanged.
- Normalize Date to YYYY-MM-DD and Punch Time to HH:MM:SS.
"""
import pandas as pd
from openpyxl import Workbook

from data_loader import (
    _adapt_attendance_dataframe,
    _normalize_attendance_date,
    _normalize_punch_time,
    load_attendance_file,
)


# -- normalizers (unit) ---------------------------------------------------

def test_normalize_date_handles_iso_dd_mm_yyyy_and_timestamp():
    assert _normalize_attendance_date("2026-04-26") == "2026-04-26"
    assert _normalize_attendance_date("26-04-2026") == "2026-04-26"
    assert _normalize_attendance_date(pd.Timestamp("2026-04-26")) == "2026-04-26"
    # Anything not a date stays a string (caller decides what to do).
    assert _normalize_attendance_date("junk") == "junk"


def test_normalize_punch_time_appends_seconds_when_missing():
    assert _normalize_punch_time("09:43:00") == "09:43:00"
    assert _normalize_punch_time("09:43") == "09:43:00"
    # datetime.time / Timestamp objects.
    from datetime import time
    assert _normalize_punch_time(time(9, 43)) == "09:43:00"


# -- _adapt_attendance_dataframe (unit) -----------------------------------

def test_adapter_passes_legacy_punch_schema_through():
    df = pd.DataFrame([
        {"Employee ID": 1, "First Name": "ALI-EMP1",
         "Date": "2026-04-26", "Punch Time": "08:00:00",
         "Punch State": "Check In"},
        {"Employee ID": 1, "First Name": "ALI-EMP1",
         "Date": "2026-04-26", "Punch Time": "17:00:00",
         "Punch State": "Check Out"},
    ])
    out = _adapt_attendance_dataframe(df)
    # Same shape, same columns, same values.
    assert list(out.columns) == list(df.columns)
    assert len(out) == 2
    assert (out["Punch State"] == ["Check In", "Check Out"]).all()


def test_adapter_converts_first_in_last_out_to_punch_events():
    df = pd.DataFrame([
        {
            "Employee ID": 4100323,
            "First Name": "MOHAMMED-EMP422",
            "Date": "26-04-2026",
            "Weekday": "Sunday",
            "First Check In": "09:43",
            "Last Check Out": "18:05",
            "Total Time": "08:22",
        },
    ])
    out = _adapt_attendance_dataframe(df)
    assert list(out.columns) == [
        "Employee ID", "First Name", "Date", "Punch Time", "Punch State",
    ]
    assert len(out) == 2  # one Check In + one Check Out
    assert out.iloc[0]["Punch State"] == "Check In"
    assert out.iloc[0]["Punch Time"] == "09:43:00"
    assert out.iloc[0]["Date"] == "2026-04-26"  # normalized to ISO
    assert out.iloc[1]["Punch State"] == "Check Out"
    assert out.iloc[1]["Punch Time"] == "18:05:00"
    assert out.iloc[1]["Date"] == "2026-04-26"


def test_adapter_skips_missing_punches_per_row():
    """No Check In -> only the Check Out is emitted, and vice versa."""
    df = pd.DataFrame([
        {
            "Employee ID": 1, "First Name": "A-EMP1",
            "Date": "27-04-2026", "Weekday": "Monday",
            "First Check In": None,         # missed
            "Last Check Out": "17:56",
            "Total Time": None,
        },
        {
            "Employee ID": 1, "First Name": "A-EMP1",
            "Date": "28-04-2026", "Weekday": "Tuesday",
            "First Check In": "09:00",
            "Last Check Out": None,         # missed
            "Total Time": None,
        },
    ])
    out = _adapt_attendance_dataframe(df)
    assert len(out) == 2  # exactly one event per source row
    assert (out["Punch State"] == ["Check Out", "Check In"]).all()
    assert out.iloc[0]["Date"] == "2026-04-27"
    assert out.iloc[1]["Date"] == "2026-04-28"


def test_adapter_rejects_unknown_schema():
    df = pd.DataFrame([{"foo": 1, "bar": 2}])
    try:
        _adapt_attendance_dataframe(df)
    except ValueError as e:
        assert "Unrecognized attendance export schema" in str(e)
    else:
        raise AssertionError("expected ValueError")


# -- end-to-end with real XLSX fixtures ----------------------------------

def _write_legacy_xlsx(path):
    """Write a 1-banner-row Transaction-template XLSX fixture."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Transaction Report"])  # banner
    ws.append(["Employee ID", "First Name", "Date",
               "Punch Time", "Punch State"])
    ws.append([1, "ALI-EMP1", "2026-04-26", "08:00:00", "Check In"])
    ws.append([1, "ALI-EMP1", "2026-04-26", "17:00:00", "Check Out"])
    wb.save(path)


def _write_first_in_last_out_xlsx(path):
    """Write a 4-banner-row First In Last Out template XLSX fixture."""
    wb = Workbook()
    ws = wb.active
    ws.append(["First In Last Out Report"])
    ws.append(["From  April 26 2026  To  May 18 2026"])
    ws.append(["Company: electron", "", "", "",
               "Generated On: Tue May 19 2026 10:25:35"])
    ws.append([])  # blank
    ws.append(["Employee ID", "First Name", "Date", "Weekday",
               "First Check In", "Last Check Out", "Total Time"])
    ws.append([4100323, "MOHAMMED-EMP422", "26-04-2026", "Sunday",
               "09:43", "18:05", "08:22"])
    wb.save(path)


def test_load_attendance_file_legacy_template(tmp_path):
    path = tmp_path / "legacy.xlsx"
    _write_legacy_xlsx(path)
    df = load_attendance_file(path)
    assert list(df.columns) == [
        "Employee ID", "First Name", "Date", "Punch Time", "Punch State",
    ]
    assert len(df) == 2
    assert df.iloc[0]["Punch State"] == "Check In"


def test_load_attendance_file_first_in_last_out_template(tmp_path):
    path = tmp_path / "fil.xlsx"
    _write_first_in_last_out_xlsx(path)
    df = load_attendance_file(path)
    assert list(df.columns) == [
        "Employee ID", "First Name", "Date", "Punch Time", "Punch State",
    ]
    assert len(df) == 2  # 1 source row -> Check In + Check Out
    assert df.iloc[0]["Date"] == "2026-04-26"
    assert df.iloc[0]["Punch Time"] == "09:43:00"
    assert df.iloc[0]["Punch State"] == "Check In"
    assert df.iloc[1]["Punch Time"] == "18:05:00"
    assert df.iloc[1]["Punch State"] == "Check Out"


def test_load_attendance_file_raises_on_missing_header(tmp_path):
    """If no header row contains Employee ID, the loader must explain
    rather than crash deep in the pipeline."""
    path = tmp_path / "broken.xlsx"
    wb = Workbook()
    ws = wb.active
    for _ in range(5):
        ws.append(["random", "banner", "stuff"])
    wb.save(path)
    try:
        load_attendance_file(path)
    except ValueError as e:
        assert "Could not find an attendance header row" in str(e)
    else:
        raise AssertionError("expected ValueError")
