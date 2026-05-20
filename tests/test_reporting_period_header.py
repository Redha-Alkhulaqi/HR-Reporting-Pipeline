"""Reporting Period banner must appear at the top of every executive
sheet in the exported workbook.

Spec:
  Row 1: 'Reporting Period: YYYY-MM-DD to YYYY-MM-DD'   (bold, size 14)
  Row 2: 'Generated On: YYYY-MM-DD HH:MM AM/PM'         (bold, size 11)
  Row 3: blank spacer
  Row 4: table header
  Row 5+: data

Sheets covered: Dashboard, Employee Summary, Employee Attendance,
Daily Attendance, Daily Trend, Department Summary, plus every audit
sheet that goes through `_build_data_sheet`.

Backward compat: when period_start / period_end are both None,
sheets keep the legacy row-1 header (no banner) so existing tests
remain valid.
"""
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from excel_exporter import (
    _PERIOD_BLOCK_ROWS,
    _build_data_sheet,
    _build_employee_attendance_sheet,
    _build_executive_employee_sheet,
    _format_period_date,
    _write_reporting_period_block,
    export_report,
)


# ---- Helper-level contracts ---------------------------------------------

def test_format_period_date_handles_strings_dates_and_none():
    assert _format_period_date("2026-05-19") == "2026-05-19"
    assert _format_period_date(pd.Timestamp("2026-04-20")) == "2026-04-20"
    assert _format_period_date(None) == ""
    assert _format_period_date(pd.NaT) == ""


def test_write_reporting_period_block_writes_3_rows_and_returns_row_4():
    wb = Workbook()
    ws = wb.active
    next_row = _write_reporting_period_block(
        ws, period_start="2026-04-20", period_end="2026-05-19", n_cols=10,
    )
    assert next_row == 4

    r1 = ws.cell(row=1, column=1).value
    r2 = ws.cell(row=2, column=1).value
    assert "Reporting Period:" in r1
    assert "2026-04-20" in r1 and "2026-05-19" in r1
    assert "Generated On:" in r2

    # Both rows must be merged across `n_cols` to act as a banner.
    merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
    assert "A1:J1" in merged_ranges
    assert "A2:J2" in merged_ranges
    # Banner cells are bold + centered.
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=1, column=1).alignment.horizontal == "center"
    assert ws.cell(row=2, column=1).font.bold is True


def test_write_reporting_period_block_handles_one_sided_bounds():
    """A pipeline run with --from only (no --to) must still produce a
    sensible banner instead of two dangling 'to ' fragments."""
    wb = Workbook()
    ws = wb.active
    _write_reporting_period_block(ws, "2026-04-20", None, n_cols=5)
    assert "from 2026-04-20" in ws.cell(row=1, column=1).value


# ---- Per-sheet contracts -------------------------------------------------

PERIOD_START = "2026-04-20"
PERIOD_END = "2026-05-19"


def _build_executive_df():
    return pd.DataFrame([{
        "Employee ID": 4195162, "First Name": "ALI",
        "No of Absence Days": 0.0,
        "No of Permission Days": 0, "No of Vacation Days": 0,
        "No of Secondment Days": 0, "Total Late (Hours)": 0.0,
        "Total Over Time (Hours) (Actual)": 0.0,
        "Total Over Time (Payable 1.5x) (Hours)": 0.0,
        "Total Early Leave (Hours)": 0.0,
        "Break Time (Hours)": 0.0,
        "Break Time (After Policy)": 0.0,
        "Friday Compensation Days": 0,
        "Friday Worked Dates": "",
    }])


def _expect_banner(ws, banner_width):
    """Row 1 / Row 2 carry the banner; table header lands at row 4."""
    assert "Reporting Period:" in (ws.cell(row=1, column=1).value or "")
    assert "Generated On:" in (ws.cell(row=2, column=1).value or "")
    # Banner spans the data width.
    merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(banner_width)
    assert f"A1:{last_col}1" in merged_ranges
    assert f"A2:{last_col}2" in merged_ranges


def test_executive_employee_sheet_carries_banner_above_header_row_4():
    wb = Workbook()
    ws = wb.active
    df = _build_executive_df()
    _build_executive_employee_sheet(
        ws, df, period_start=PERIOD_START, period_end=PERIOD_END,
    )
    _expect_banner(ws, banner_width=len(df.columns))
    # Table header lands at row 4 (= _PERIOD_BLOCK_ROWS + 1).
    assert ws.cell(row=4, column=1).value == "Employee ID"
    assert ws.cell(row=4, column=2).value == "First Name"
    # Data starts at row 5; Employee ID still stored as TEXT.
    assert ws.cell(row=5, column=1).value == "4195162"
    assert ws.cell(row=5, column=1).number_format == "@"
    # Freeze pane was lifted to A5 (was A2 pre-banner).
    assert ws.freeze_panes == "A5"


def test_data_sheet_carries_banner_above_header_row_4():
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame([
        {"Employee ID": 100, "Status": "Normal"},
        {"Employee ID": 101, "Status": "Late"},
    ])
    _build_data_sheet(ws, df, period_start=PERIOD_START, period_end=PERIOD_END)
    _expect_banner(ws, banner_width=2)
    assert ws.cell(row=4, column=1).value == "Employee ID"
    assert ws.cell(row=5, column=1).value == "100"
    assert ws.cell(row=6, column=1).value == "101"
    assert ws.freeze_panes == "A5"


def test_data_sheet_without_period_is_unchanged_legacy_layout():
    """Backward compat: omitting both period bounds keeps the historical
    row-1-header layout used by older callers / tests."""
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame([{"Employee ID": 7, "Status": "Normal"}])
    _build_data_sheet(ws, df)  # no period_*
    assert ws.cell(row=1, column=1).value == "Employee ID"
    assert ws.cell(row=2, column=1).value == "7"
    assert ws.freeze_panes == "A2"


def test_employee_attendance_sheet_carries_banner():
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame([{
        "Raw Employee ID": 4195162,
        "Canonical Employee Name": "ALI",
        "Canonical Employee ID": 4195162,
        "Date": "2026-05-03", "Weekday": "Sun",
        "Shift 1 Check-In": "08:00", "Shift 1 Check-Out": "17:00",
        "Shift 2 Check-In": "", "Shift 2 Check-Out": "",
        "Total Time": "09:00", "Source / Notes": "BioTime",
    }])
    _build_employee_attendance_sheet(
        ws, df, period_start=PERIOD_START, period_end=PERIOD_END,
    )
    _expect_banner(ws, banner_width=len(df.columns))
    # Original freeze of "F2" shifts to "F5".
    assert ws.freeze_panes == "F5"
    assert ws.cell(row=4, column=1).value == "Raw Employee ID"
    # Raw + Canonical Employee ID remain TEXT-formatted.
    assert ws.cell(row=5, column=1).value == "4195162"
    assert ws.cell(row=5, column=1).number_format == "@"


# ---- End-to-end ----------------------------------------------------------

def _build_minimal_summary():
    executive_summary = _build_executive_df()
    status_summary = pd.DataFrame([
        {"attendance_status": "Normal", "count": 1},
    ])
    daily = pd.DataFrame([{
        "Employee ID": 4195162, "First Name": "ALI",
        "Date": "2026-05-03",
        "Check In": "08:00:00", "Check Out": "17:00:00",
        "Shift Start": "08:00:00", "Shift End": "17:00:00",
        "worked_minutes": 540, "scheduled_minutes": 540,
        "is_late": False, "unexcused_delay_minutes": 0,
        "overtime_minutes": 0, "early_leave_minutes": 0,
        "early_leave_anomaly": False, "missing_check_out": False,
        "attendance_status": "Normal", "is_excluded": False,
        "overtime_status": "Normal", "early_leave_status": "Normal",
    }])
    department_summary = pd.DataFrame([
        {"Department": "Operations", "On Time": 5, "Late": 1},
    ])
    daily_trend = pd.DataFrame([
        {"Date": "2026-05-03", "total_records": 1, "late_cases": 0},
    ])
    summary = {
        "executive_employee_summary": executive_summary,
        "status_summary": status_summary,
        "daily_trend": daily_trend,
        "department_summary": department_summary,
        "late_cases": 0, "total_late_minutes": 0,
        "approved_excuse_cases": 0, "leave_cases": 0,
        "missing_schedule_cases": 0, "missing_check_out_cases": 0,
        "high_risk_employees": 0, "excluded_employee_count": 0,
        "total_deduction_capped": 0, "overtime_cases": 0,
        "total_overtime_hours": 0, "overtime_multiplier": 1.5,
        "total_overtime_payable_hours": 0, "early_leave_cases": 0,
        "total_early_leave_minutes": 0, "early_leave_anomaly_cases": 0,
        "total_break_count": 0, "total_break_minutes": 0,
        "employees_with_breaks": 0, "incomplete_break_records": 0,
        "employee_id_aliases_used": 0, "employee_id_alias_records_mapped": 0,
        "data_quality_score": 100, "reporting_population": 1,
    }
    return summary, daily


def test_export_report_renders_banner_on_dashboard_and_data_sheets(
    tmp_path, monkeypatch,
):
    monkeypatch.setattr("excel_exporter.REPORT_OUTPUT_DIR", tmp_path)
    summary, daily = _build_minimal_summary()
    export_report(
        summary, daily,
        period_start=PERIOD_START, period_end=PERIOD_END,
    )
    written = list(tmp_path.rglob("hr_report_*.xlsx"))
    assert written, "expected a workbook in the temp output dir"
    wb = load_workbook(written[0])

    # Dashboard: title at row 1, then Reporting Period at row 2,
    # Generated On at row 3, KPI header at row 5.
    dash = wb["Dashboard"]
    assert dash.cell(row=1, column=1).value == "HR Reporting Dashboard"
    assert "Reporting Period:" in (dash.cell(row=2, column=1).value or "")
    assert PERIOD_START in dash.cell(row=2, column=1).value
    assert PERIOD_END in dash.cell(row=2, column=1).value
    assert "Generated On:" in (dash.cell(row=3, column=1).value or "")
    assert dash.cell(row=5, column=1).value == "Metric"
    assert dash.cell(row=5, column=2).value == "Value"

    # Every data sheet must carry the banner in rows 1-2 and have its
    # header at row 4. Use a representative subset.
    for sheet_name in (
        "Employee Summary", "Daily Attendance", "Daily Trend",
        "Department Summary",
    ):
        ws = wb[sheet_name]
        assert "Reporting Period:" in (ws.cell(row=1, column=1).value or ""), (
            f"{sheet_name} row 1 missing Reporting Period banner"
        )
        assert "Generated On:" in (ws.cell(row=2, column=1).value or ""), (
            f"{sheet_name} row 2 missing Generated On line"
        )
        # Row 4 holds the table header; it must be non-empty.
        assert ws.cell(row=4, column=1).value not in (None, ""), (
            f"{sheet_name} row 4 (expected table header) is empty"
        )


def test_export_report_without_period_keeps_legacy_layout(
    tmp_path, monkeypatch,
):
    """When period_start / period_end are omitted, the executive sheets
    keep the historical row-1 header layout."""
    monkeypatch.setattr("excel_exporter.REPORT_OUTPUT_DIR", tmp_path)
    summary, daily = _build_minimal_summary()
    export_report(summary, daily)  # no period kwargs

    written = list(tmp_path.rglob("hr_report_*.xlsx"))
    wb = load_workbook(written[0])
    es = wb["Employee Summary"]
    assert es.cell(row=1, column=1).value == "Employee ID"  # no banner
    dash = wb["Dashboard"]
    # Legacy "Generated YYYY-MM-DD HH:MM" subtitle, NOT the new banner.
    assert dash.cell(row=2, column=1).value.startswith("Generated ")
    assert "Reporting Period:" not in (dash.cell(row=2, column=1).value or "")
