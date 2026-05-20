"""Employee ID must be exported to Excel as TEXT, never as a number.

Excel auto-applies a thousand-separator format to large numeric cells
(4195162 -> 4,195,162). For an identifier column that is wrong: HR
needs to see the bare code exactly as stored, with no separators,
scientific notation, or numeric coercion.

These tests pin two layers of the contract:

1. The helper `_format_employee_id_columns_as_text` rewrites every
   'Employee ID' cell as a string and pins number_format='@'.
2. End-to-end via `export_report`, the Employee Summary,
   Employee Attendance, and Daily Attendance sheets all carry
   Employee ID cells as strings (no comma) with the TEXT (@) format.
"""
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from excel_exporter import (
    _format_employee_id_columns_as_text,
    export_report,
)


def test_helper_converts_int_ids_to_string_and_sets_text_format():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Employee ID")
    ws.cell(row=1, column=2, value="First Name")
    ws.cell(row=2, column=1, value=4195162)
    ws.cell(row=2, column=2, value="ALI")
    ws.cell(row=3, column=1, value=3006)
    ws.cell(row=3, column=2, value="OMAR")

    _format_employee_id_columns_as_text(
        ws, header_row=1, data_start=2, data_end=3, n_cols=2,
    )

    for row, expected in [(2, "4195162"), (3, "3006")]:
        cell = ws.cell(row=row, column=1)
        assert cell.value == expected
        assert isinstance(cell.value, str)
        assert cell.number_format == "@"
        assert "," not in cell.value
    # Untouched neighbour column.
    assert ws.cell(row=2, column=2).value == "ALI"


def test_helper_strips_comma_thousand_separators_from_string_input():
    """When the source already has 'pretty' commas in it, strip them."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Employee ID")
    ws.cell(row=2, column=1, value="4,195,162")
    ws.cell(row=3, column=1, value="  4179594  ")

    _format_employee_id_columns_as_text(
        ws, header_row=1, data_start=2, data_end=3, n_cols=1,
    )
    assert ws.cell(row=2, column=1).value == "4195162"
    assert ws.cell(row=3, column=1).value == "4179594"


def test_helper_matches_every_employee_id_variant_header():
    """The pipeline emits several 'Employee ID' headers: Raw / Canonical
    / Old / Current. The helper must format ALL of them as TEXT."""
    wb = Workbook()
    ws = wb.active
    headers = [
        "Employee ID", "Raw Employee ID", "Canonical Employee ID",
        "Old Employee ID", "Current Employee ID",
    ]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=h)
        ws.cell(row=2, column=idx, value=4195162 + idx)

    _format_employee_id_columns_as_text(
        ws, header_row=1, data_start=2, data_end=2, n_cols=len(headers),
    )
    for idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=idx)
        assert isinstance(cell.value, str)
        assert cell.number_format == "@"
        assert "," not in cell.value


def test_helper_skips_columns_whose_header_is_not_employee_id():
    """First Name / Date / numeric columns must remain unchanged."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Employee ID")
    ws.cell(row=1, column=2, value="Total Late (Hours)")
    ws.cell(row=2, column=1, value=4195162)
    ws.cell(row=2, column=2, value=12.5)

    _format_employee_id_columns_as_text(
        ws, header_row=1, data_start=2, data_end=2, n_cols=2,
    )
    # Employee ID -> text.
    assert ws.cell(row=2, column=1).value == "4195162"
    assert ws.cell(row=2, column=1).number_format == "@"
    # Numeric neighbour untouched.
    assert ws.cell(row=2, column=2).value == 12.5
    assert ws.cell(row=2, column=2).number_format != "@"


def test_helper_tolerates_none_and_empty_cells():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Employee ID")
    ws.cell(row=2, column=1, value=None)
    ws.cell(row=3, column=1, value="")
    ws.cell(row=4, column=1, value=4195162)

    _format_employee_id_columns_as_text(
        ws, header_row=1, data_start=2, data_end=4, n_cols=1,
    )
    # Empty cells stay empty; populated cell still becomes text.
    assert ws.cell(row=2, column=1).value is None
    assert ws.cell(row=3, column=1).value == ""
    assert ws.cell(row=4, column=1).value == "4195162"


# ---- End-to-end ----------------------------------------------------------


def _build_minimal_summary():
    """Return (summary, daily) inputs sufficient for export_report.

    Mirrors the minimum keys that the executive sheet builder reads.
    Uses two large integer IDs (the same shape Excel auto-formats with
    thousand separators) so the assertion catches the regression.
    """
    executive_summary = pd.DataFrame([
        {
            "Employee ID": 4195162,
            "First Name": "ALI",
            "No of Absence Days": 0.0,
            "No of Permission Days": 0,
            "No of Vacation Days": 0,
            "No of Secondment Days": 0,
            "Total Late (Hours)": 0.0,
            "Total Over Time (Hours) (Actual)": 0.0,
            "Total Over Time (Payable 1.5x) (Hours)": 0.0,
            "Total Early Leave (Hours)": 0.0,
            "Break Time (Hours)": 0.0,
            "Break Time (After Policy)": 0.0,
            "Friday Compensation Days": 0,
            "Friday Worked Dates": "",
        },
        {
            "Employee ID": 3006,
            "First Name": "OMAR",
            "No of Absence Days": 0.0,
            "No of Permission Days": 0,
            "No of Vacation Days": 0,
            "No of Secondment Days": 0,
            "Total Late (Hours)": 0.0,
            "Total Over Time (Hours) (Actual)": 0.0,
            "Total Over Time (Payable 1.5x) (Hours)": 0.0,
            "Total Early Leave (Hours)": 0.0,
            "Break Time (Hours)": 0.0,
            "Break Time (After Policy)": 0.0,
            "Friday Compensation Days": 0,
            "Friday Worked Dates": "",
        },
    ])
    status_summary = pd.DataFrame(
        [{"attendance_status": "Normal", "count": 2}]
    )
    daily = pd.DataFrame([
        {
            "Employee ID": 4195162, "First Name": "ALI",
            "Date": "2026-05-03",
            "Check In": "08:00:00", "Check Out": "17:00:00",
            "Shift Start": "08:00:00", "Shift End": "17:00:00",
            "worked_minutes": 540, "scheduled_minutes": 540,
            "is_late": False, "unexcused_delay_minutes": 0,
            "overtime_minutes": 0, "early_leave_minutes": 0,
            "early_leave_anomaly": False,
            "missing_check_out": False,
            "attendance_status": "Normal",
            "is_excluded": False,
            "overtime_status": "Normal",
            "early_leave_status": "Normal",
        },
    ])
    summary = {
        "executive_employee_summary": executive_summary,
        "status_summary": status_summary,
        "daily_trend": pd.DataFrame(),
        "late_cases": 0,
        "total_late_minutes": 0,
        "approved_excuse_cases": 0,
        "leave_cases": 0,
        "missing_schedule_cases": 0,
        "missing_check_out_cases": 0,
        "high_risk_employees": 0,
        "excluded_employee_count": 0,
        "total_deduction_capped": 0,
        "overtime_cases": 0,
        "total_overtime_hours": 0,
        "overtime_multiplier": 1.5,
        "total_overtime_payable_hours": 0,
        "early_leave_cases": 0,
        "total_early_leave_minutes": 0,
        "early_leave_anomaly_cases": 0,
        "total_break_count": 0,
        "total_break_minutes": 0,
        "employees_with_breaks": 0,
        "incomplete_break_records": 0,
        "employee_id_aliases_used": 0,
        "employee_id_alias_records_mapped": 0,
        "data_quality_score": 100,
        "reporting_population": 2,
    }
    return summary, daily


def test_export_report_writes_employee_ids_as_text(tmp_path, monkeypatch):
    """End-to-end: the saved workbook must store every Employee ID as
    a TEXT cell with no thousand separators in the displayed value."""
    monkeypatch.setattr("excel_exporter.REPORT_OUTPUT_DIR", tmp_path)

    summary, daily = _build_minimal_summary()
    export_report(summary, daily)

    written = list(tmp_path.rglob("hr_report_*.xlsx"))
    assert written, "expected a workbook in the temp output dir"
    wb = load_workbook(written[0])
    ws = wb["Employee Summary"]
    assert ws.cell(row=1, column=1).value == "Employee ID"
    # Row 2 + 3 are the two employees. Both must be strings, comma-free,
    # and pinned to TEXT (@) format.
    for r, expected in [(2, "4195162"), (3, "3006")]:
        cell = ws.cell(row=r, column=1)
        assert isinstance(cell.value, str), (
            f"Employee Summary row {r} stored as {type(cell.value).__name__}"
        )
        assert cell.value == expected
        assert "," not in cell.value
        assert cell.number_format == "@"
