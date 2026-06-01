"""Executive chart-polish contract for Dashboard pie + line.

The reference UI/UX image (see review thread) demands:
  Pie chart:
    - subtitle below the main title
    - percentage shown directly on slices
    - legend on the right, not overlapping
    - labels positioned by Excel's bestFit so small slices stay readable
  Line chart:
    - subtitle below the main title
    - X-axis title "Date"
    - Y-axis title "Late Cases"
    - major Y-axis gridlines so values are readable at a glance
    - circle markers on every data point
    - rotated X-axis date labels so they don't overlap

These tests pin those properties on the chart objects themselves, so a
future tweak to `_pie_chart` / `_line_chart` that drops one of the
executive defaults fails CI fast.
"""
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import Reference

from excel_exporter import (
    _chart_title,
    _line_chart,
    _pie_chart,
)


# ---- Title + subtitle helper --------------------------------------------

def test_chart_title_builds_two_paragraphs_when_subtitle_supplied():
    title = _chart_title("Attendance", "(Distribution %)")
    # Title.tx.rich.p has the paragraph list.
    paragraphs = title.tx.rich.p
    assert len(paragraphs) == 2
    # Title paragraph
    assert paragraphs[0].r[0].t == "Attendance"
    # Subtitle paragraph
    assert paragraphs[1].r[0].t == "(Distribution %)"
    # Subtitle font is smaller + italic
    assert paragraphs[1].r[0].rPr.sz == 1000
    assert paragraphs[1].r[0].rPr.i is True
    # Title is bold, larger
    assert paragraphs[0].r[0].rPr.b is True
    assert paragraphs[0].r[0].rPr.sz == 1400
    assert title.overlay is False


def test_chart_title_without_subtitle_has_one_paragraph():
    title = _chart_title("Late Cases")
    paragraphs = title.tx.rich.p
    assert len(paragraphs) == 1
    assert paragraphs[0].r[0].t == "Late Cases"


# ---- Pie chart polish ---------------------------------------------------

def _build_pie():
    wb = Workbook()
    ws = wb.active
    ws.append(["Status", "Count"])
    ws.append(["On Time", 78])
    ws.append(["Late", 19])
    ws.append(["Approved Excuse", 3])
    labels = Reference(ws, min_col=1, min_row=2, max_row=4)
    values = Reference(ws, min_col=2, min_row=2, max_row=4)
    return _pie_chart(
        "Attendance Status Breakdown", labels, values,
        subtitle="(Distribution %)",
    )


def test_pie_chart_title_has_subtitle():
    chart = _build_pie()
    paragraphs = chart.title.tx.rich.p
    assert len(paragraphs) == 2
    assert "Distribution" in paragraphs[1].r[0].t


def test_pie_data_labels_show_percent_only_not_category_name():
    """Category names live in the legend so the labels on the slices
    can stay short and readable on the small 3%/0% slices."""
    chart = _build_pie()
    dl = chart.dataLabels
    assert dl.showPercent is True
    assert dl.showCatName is False
    assert dl.showVal is False
    assert dl.showSerName is False


def test_pie_data_labels_use_best_fit_positioning():
    """openpyxl exposes the position attribute as `dLblPos`. 'bestFit'
    lets Excel push tight labels outside the slice with a leader line."""
    chart = _build_pie()
    assert chart.dataLabels.dLblPos == "bestFit"


def test_pie_legend_on_right_and_not_overlaying():
    chart = _build_pie()
    assert chart.legend is not None
    assert chart.legend.position == "r"
    assert chart.legend.overlay is False


# ---- Line chart polish --------------------------------------------------

def _build_line():
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Late Cases"])
    for d, v in [
        ("2026-05-01", 2), ("2026-05-02", 15),
        ("2026-05-03", 30), ("2026-05-06", 50),
    ]:
        ws.append([d, v])
    labels = Reference(ws, min_col=1, min_row=2, max_row=5)
    values = Reference(ws, min_col=2, min_row=1, max_row=5)
    return _line_chart(
        "Daily Late Trend", labels, values,
        subtitle="(Number of Late Cases)",
        x_axis_title="Date",
        y_axis_title="Late Cases",
    )


def test_line_chart_title_has_subtitle():
    chart = _build_line()
    paragraphs = chart.title.tx.rich.p
    assert len(paragraphs) == 2
    assert "Number of Late Cases" in paragraphs[1].r[0].t


def test_line_chart_axis_titles_set():
    chart = _build_line()
    assert chart.x_axis.title is not None
    assert chart.y_axis.title is not None


def test_line_chart_y_axis_major_gridlines_enabled():
    chart = _build_line()
    assert chart.y_axis.majorGridlines is not None


def test_line_chart_points_have_circle_markers():
    chart = _build_line()
    assert chart.series, "line chart should have at least one series"
    for s in chart.series:
        assert s.marker is not None
        assert s.marker.symbol == "circle"
        assert s.marker.size == 7


def test_line_chart_x_axis_labels_rotated_negative_45_degrees():
    """Rotation is stored in 1/60000 of a degree. -45deg = -2_700_000."""
    chart = _build_line()
    assert chart.x_axis.txPr is not None
    assert chart.x_axis.txPr.bodyPr is not None
    assert chart.x_axis.txPr.bodyPr.rot == -2_700_000


def test_line_chart_has_no_legend_for_single_series():
    """Single-series line; the title carries the meaning, no legend
    needed (and a legend would just clutter the plot area)."""
    chart = _build_line()
    assert chart.legend is None


# ---- Bar chart polish ---------------------------------------------------

from excel_exporter import _bar_chart


def _build_bar():
    wb = Workbook()
    ws = wb.active
    ws.append(["Employee", "Late Hours"])
    for name, hrs in [
        ("Ali Al-Balawi", 13.6),
        ("Abdullah Al-Anzi", 12.45),
        ("Khalid Al-Subaie", 11.2),
    ]:
        ws.append([name, hrs])
    labels = Reference(ws, min_col=1, min_row=2, max_row=4)
    values = Reference(ws, min_col=2, min_row=2, max_row=4)
    return _bar_chart(
        "Top Late Employees", labels, values, horizontal=True,
        subtitle="By Total Late Hours (Hours)",
        x_axis_title="Late Hours (Hours)",
    )


def test_bar_chart_title_has_subtitle():
    chart = _build_bar()
    paragraphs = chart.title.tx.rich.p
    assert len(paragraphs) == 2
    assert "Total Late Hours" in paragraphs[1].r[0].t


def test_bar_chart_data_labels_show_value_at_bar_outside_end():
    """Every bar must have its numeric value rendered at the outside
    end of the bar -- the #1 readability fix the user requested."""
    chart = _build_bar()
    dl = chart.dataLabels
    assert dl is not None, "bar chart must have data labels"
    assert dl.showVal is True
    assert dl.showCatName is False
    assert dl.showSerName is False
    assert dl.dLblPos == "outEnd"


def test_bar_chart_axis_title_set():
    chart = _build_bar()
    assert chart.x_axis.title is not None


def test_bar_chart_category_labels_pinned_to_low_side():
    """Category labels (employee names) must render on the left of
    the horizontal bar chart. Without `tickLblPos="low"` Excel can
    flip them to the bar's far end when bars become narrow."""
    chart = _build_bar()
    assert chart.y_axis.tickLblPos == "low"


def test_bar_chart_axis_labels_use_9pt_font():
    chart = _build_bar()
    assert chart.y_axis.txPr is not None
    # The defRPr on the first paragraph carries the size (1/100 pt).
    sz = chart.y_axis.txPr.p[0].pPr.defRPr.sz
    assert sz == 900


def test_bar_chart_has_no_legend_for_single_series():
    chart = _build_bar()
    assert chart.legend is None


# ---- "How to read" annotation box --------------------------------------

from excel_exporter import _write_chart_annotation


# ---- Top Overtime accuracy regression -----------------------------------

def test_top_overtime_chart_plots_payable_hours_not_minutes(
    tmp_path, monkeypatch,
):
    """Top Overtime chart MUST reference the `total_overtime_payable_hours`
    column (not `total_overtime_minutes`).

    Bug context: the chart's title/subtitle/X-axis title all promise
    "Hours (Payable 1.5x)" but the Reference used to point at
    `total_overtime_minutes`, producing a bar height that was 60x the
    value it claimed. This test pins the source column AND the
    rendered cell values so the chart can't silently drift back to
    minutes.
    """
    from excel_exporter import export_report
    import pandas as pd

    monkeypatch.setattr("excel_exporter.REPORT_OUTPUT_DIR", tmp_path)

    # Build a top_overtime_employees table that carries BOTH the
    # minutes and the payable-hours columns -- mirrors what the live
    # pipeline emits when the 1.5x multiplier is active.
    top_overtime = pd.DataFrame([
        {
            "Employee ID": 4184410,
            "First Name": "QASEM-EMP416",
            "overtime_cases": 22, "total_overtime_minutes": 3242,
            "avg_overtime_minutes": 147, "total_overtime_hours": 54.0,
            "total_overtime_payable_minutes": 4863,
            "total_overtime_payable_hours": 81.0,
        },
        {
            "Employee ID": 4125249,
            "First Name": "ALAA-EMP410",
            "overtime_cases": 18, "total_overtime_minutes": 2406,
            "avg_overtime_minutes": 134, "total_overtime_hours": 40.1,
            "total_overtime_payable_minutes": 3609,
            "total_overtime_payable_hours": 60.15,
        },
    ])
    executive = pd.DataFrame([{
        "Employee ID": 4184410, "First Name": "QASEM-EMP416",
        "No of Absence Days": 0.0, "No of Permission Days": 0,
        "No of Vacation Days": 0, "No of Secondment Days": 0,
        "Total Late (Hours)": 0.0,
        "Total Over Time (Hours) (Actual)": 54.0,
        "Total Over Time (Payable 1.5x) (Hours)": 81.0,
        "Total Early Leave (Hours)": 0.0,
        "Break Time (Hours)": 0.0, "Break Time (After Policy)": 0.0,
        "Friday Compensation Days": 0, "Friday Worked Dates": "",
    }])
    summary = {
        "executive_employee_summary": executive,
        "status_summary": pd.DataFrame(
            [{"attendance_status": "Normal", "count": 1}]
        ),
        "daily_trend": pd.DataFrame(),
        "top_overtime_employees": top_overtime,
        "late_cases": 0, "total_late_minutes": 0,
        "approved_excuse_cases": 0, "leave_cases": 0,
        "missing_schedule_cases": 0, "missing_check_out_cases": 0,
        "high_risk_employees": 0, "excluded_employee_count": 0,
        "total_deduction_capped": 0, "overtime_cases": 2,
        "total_overtime_hours": 94.1, "overtime_multiplier": 1.5,
        "total_overtime_payable_hours": 141.15,
        "early_leave_cases": 0, "total_early_leave_minutes": 0,
        "early_leave_anomaly_cases": 0, "total_break_count": 0,
        "total_break_minutes": 0, "employees_with_breaks": 0,
        "incomplete_break_records": 0, "employee_id_aliases_used": 0,
        "employee_id_alias_records_mapped": 0, "data_quality_score": 100,
        "reporting_population": 1,
    }
    daily = pd.DataFrame([{
        "Employee ID": 4184410, "First Name": "QASEM-EMP416",
        "Date": "2026-05-03",
        "Check In": "08:00:00", "Check Out": "20:00:00",
        "Shift Start": "08:00:00", "Shift End": "17:00:00",
        "worked_minutes": 720, "scheduled_minutes": 540,
        "is_late": False, "unexcused_delay_minutes": 0,
        "overtime_minutes": 180, "early_leave_minutes": 0,
        "early_leave_anomaly": False, "missing_check_out": False,
        "attendance_status": "Normal", "is_excluded": False,
        "overtime_status": "Normal", "early_leave_status": "Normal",
    }])

    export_report(summary, daily,
                  period_start="2026-05-01", period_end="2026-05-31")
    written = list(tmp_path.rglob("hr_report_*.xlsx"))
    assert written
    wb = load_workbook(written[0])

    # 1. The Dashboard Data hidden sheet must expose
    #    `total_overtime_payable_hours` as one of the Top Overtime
    #    backing-table columns.
    hd = wb["Dashboard Data"]
    section_row = next(
        r for r in range(1, hd.max_row + 1)
        if hd.cell(row=r, column=1).value == "Top Overtime Employees"
    )
    header_row = section_row + 1
    headers = [hd.cell(row=header_row, column=c).value
               for c in range(1, 7)]
    assert "total_overtime_payable_hours" in headers, (
        f"backing table missing payable_hours column: {headers}"
    )
    payable_hours_col = headers.index("total_overtime_payable_hours") + 1
    assert payable_hours_col == 6, (
        f"expected payable_hours at col 6, got {payable_hours_col}"
    )

    # 2. The Dashboard's Top Overtime chart must reference col 6
    #    (payable_hours), not col 3 (minutes).
    dash = wb["Dashboard"]
    top_ot_chart = None
    for chart in dash._charts:
        # Title is RichText; concatenate all runs to one searchable string.
        try:
            paragraphs = chart.title.tx.rich.p
            text = "".join(
                run.t for para in paragraphs for run in para.r
            )
        except AttributeError:
            text = ""
        if "Top Overtime Employees" in text:
            top_ot_chart = chart
            break
    assert top_ot_chart is not None, "Top Overtime chart not found"
    # The value series sits at index 0 (single-series bar). Inspect its
    # numRef formula to confirm it points at col F = column 6 of
    # 'Dashboard Data'.
    series = top_ot_chart.series[0]
    val_ref = series.val.numRef.f
    assert "$F$" in val_ref, (
        f"Top Overtime chart value Reference points at {val_ref!r}; "
        "expected col F (= total_overtime_payable_hours)."
    )
    assert "$C$" not in val_ref, (
        f"Top Overtime chart value Reference still points at col C "
        f"(minutes) -- the unit bug regressed. Reference: {val_ref!r}"
    )

    # 3. The data cells in col 6 carry the payable HOURS values
    #    (1 decimal), not minutes.
    data_row_1 = header_row + 1
    payable_cell = hd.cell(row=data_row_1, column=6)
    assert float(payable_cell.value) == 81.0
    assert payable_cell.number_format == "0.0"
    # Cross-check: col 3 still carries minutes (so neither column was
    # mislabelled by the fix). Col 3 value should be 3242, not 81.
    minutes_cell = hd.cell(row=data_row_1, column=3)
    assert float(minutes_cell.value) == 3242
    assert minutes_cell.number_format == "#,##0"


def test_top_late_chart_still_uses_late_hours_column(tmp_path, monkeypatch):
    """Cross-check: the Top Late chart references the Employee Summary
    sheet's col 7 (= "Total Late (Hours)"). The accuracy audit found
    this chart to be CORRECT pre-fix; this test pins it that way."""
    from excel_exporter import export_report
    import pandas as pd

    monkeypatch.setattr("excel_exporter.REPORT_OUTPUT_DIR", tmp_path)
    executive = pd.DataFrame([{
        "Employee ID": 100 + i, "First Name": f"EMP{i}",
        "No of Absence Days": 0.0, "No of Permission Days": 0,
        "No of Vacation Days": 0, "No of Secondment Days": 0,
        "Total Late (Hours)": 10.0 - i,
        "Total Over Time (Hours) (Actual)": 0.0,
        "Total Over Time (Payable 1.5x) (Hours)": 0.0,
        "Total Early Leave (Hours)": 0.0,
        "Break Time (Hours)": 0.0, "Break Time (After Policy)": 0.0,
        "Friday Compensation Days": 0, "Friday Worked Dates": "",
    } for i in range(3)])
    summary = {
        "executive_employee_summary": executive,
        "status_summary": pd.DataFrame(
            [{"attendance_status": "Normal", "count": 3}]
        ),
        "daily_trend": pd.DataFrame(),
        "late_cases": 0, "total_late_minutes": 0,
        "approved_excuse_cases": 0, "leave_cases": 0,
        "missing_schedule_cases": 0, "missing_check_out_cases": 0,
        "high_risk_employees": 0, "excluded_employee_count": 0,
        "total_deduction_capped": 0, "overtime_cases": 0,
        "total_overtime_hours": 0, "overtime_multiplier": 1.5,
        "total_overtime_payable_hours": 0, "early_leave_cases": 0,
        "total_early_leave_minutes": 0, "early_leave_anomaly_cases": 0,
        "total_break_count": 0, "total_break_minutes": 0,
        "employees_with_breaks": 0, "incomplete_break_records": 0,
        "employee_id_aliases_used": 0,
        "employee_id_alias_records_mapped": 0, "data_quality_score": 100,
        "reporting_population": 3,
    }
    daily = pd.DataFrame([{
        "Employee ID": 100, "First Name": "EMP0",
        "Date": "2026-05-03", "Check In": "08:00:00", "Check Out": "17:00:00",
        "Shift Start": "08:00:00", "Shift End": "17:00:00",
        "worked_minutes": 540, "scheduled_minutes": 540,
        "is_late": False, "unexcused_delay_minutes": 0,
        "overtime_minutes": 0, "early_leave_minutes": 0,
        "early_leave_anomaly": False, "missing_check_out": False,
        "attendance_status": "Normal", "is_excluded": False,
        "overtime_status": "Normal", "early_leave_status": "Normal",
    }])
    export_report(summary, daily,
                  period_start="2026-05-01", period_end="2026-05-31")
    written = list(tmp_path.rglob("hr_report_*.xlsx"))
    wb = load_workbook(written[0])
    dash = wb["Dashboard"]
    top_late_chart = None
    for chart in dash._charts:
        try:
            paragraphs = chart.title.tx.rich.p
            text = "".join(run.t for p in paragraphs for run in p.r)
        except AttributeError:
            text = ""
        if "Top Late Employees" in text:
            top_late_chart = chart
            break
    assert top_late_chart is not None
    val_ref = top_late_chart.series[0].val.numRef.f
    # Employee Summary col 7 = "Total Late (Hours)".
    assert "$G$" in val_ref, (
        f"Top Late chart value Reference points at {val_ref!r}; "
        "expected col G (= Total Late (Hours))."
    )


def test_how_to_read_annotation_writes_title_and_body_rows():
    wb = Workbook()
    ws = wb.active
    _write_chart_annotation(
        ws, top_row=10, col_left="A", col_right="F",
        chart_label="Top Late Employees",
        body="Shows the 10 employees with the highest total late hours.",
        accent_primary="C00000",
    )
    title_cell = ws["A10"]
    body_cell = ws["A11"]
    assert title_cell.value == "How to read - Top Late Employees"
    assert title_cell.font.bold is True
    # White text on the accent fill.
    assert title_cell.font.color.value == "00FFFFFF"
    assert title_cell.fill.fgColor.value.endswith("C00000")
    # Body row holds the explanation, italic + wrapped.
    assert "10 employees" in body_cell.value
    assert body_cell.font.italic is True
    assert body_cell.alignment.wrap_text is True
    # Both rows merged across the requested span (A:F = 6 cols).
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "A10:F10" in merged
    assert "A11:F11" in merged
