"""Phase-1 executive Dashboard layout contract.

The Dashboard is split into 4 visible vertical zones (see
`src/excel_exporter.py` -- `_build_dashboard` docstring):

    Zone 1: Header                    rows 1-5
    Zone 2: 4 hero KPI cards          rows 6-10  (cols A:L within A:M)
    Zone 3: Charts                    rows 11-?
    Zone 4: Detail KPI sections       rows 82-?

Backing tables that drive the charts live on a HIDDEN helper sheet
named "Dashboard Data" -- they used to share Zone 5 on the visible
Dashboard sheet, but the polish pass moved them off so the Dashboard
viewport stays focused on KPIs + charts + sections.

These tests pin the layout so future tweaks to the Dashboard don't
accidentally collapse one zone into another. They do NOT assert on
business numbers -- those live in the per-metric calculator tests.
"""
import pandas as pd
import pytest
from openpyxl import load_workbook

from excel_exporter import (
    _DASH_CARDS_ROW,
    _DASH_CARD_HEIGHT,
    _DASH_CHARTS_ROW,
    _DASH_DETAIL_SECTIONS_ROW,
    _DASH_HELPER_SHEET_NAME,
    _DASH_SECTION_COLORS,
    export_report,
)


# ---- Minimal pipeline output fixture ------------------------------------

PERIOD_START = "2026-04-20"
PERIOD_END = "2026-05-19"


def _executive_df():
    return pd.DataFrame([{
        "Employee ID": 4195162, "First Name": "ALI",
        "No of Absence Days": 0.0,
        "No of Permission Days": 0, "No of Vacation Days": 0,
        "No of Secondment Days": 0, "Total Late (Hours)": 0.0,
        "Total Over Time (Hours) (Actual)": 0.0,
        "Total Over Time (Payable 1.5x) (Hours)": 0.0,
        "Total Early Leave (Hours)": 0.0,
        "Break Time (Hours)": 0.0, "Break Time (After Policy)": 0.0,
        "Friday Compensation Days": 0, "Friday Worked Dates": "",
    }])


def _minimal_summary_and_daily():
    daily = pd.DataFrame([{
        "Employee ID": 4195162, "First Name": "ALI",
        "Date": "2026-05-03",
        "Check In": "08:00:00", "Check Out": "17:00:00",
        "Shift Start": "08:00:00", "Shift End": "17:00:00",
        "worked_minutes": 540, "scheduled_minutes": 540,
        "is_late": False, "unexcused_delay_minutes": 0,
        "overtime_minutes": 0, "early_leave_minutes": 0,
        "early_leave_anomaly": False, "missing_check_out": False,
        "attendance_status": "Normal", "is_excluded": False,
        "overtime_status": "Normal", "early_leave_status": "Normal",
    }])
    summary = {
        "executive_employee_summary": _executive_df(),
        "status_summary": pd.DataFrame([
            {"attendance_status": "Normal", "count": 1},
        ]),
        "daily_trend": pd.DataFrame([
            {"Date": "2026-05-03", "total_records": 1, "late_cases": 0},
        ]),
        "reporting_population": 43,
        "total_employees": 43,
        "late_cases": 151,
        "total_late_minutes": 320,
        "approved_excuse_cases": 21,
        "leave_cases": 1,
        "missing_schedule_cases": 0,
        "missing_check_out_cases": 48,
        "high_risk_employees": 4,
        "excluded_employee_count": 2,
        "total_deduction_capped": 1200.5,
        "overtime_cases": 12,
        "total_overtime_hours": 42.0,
        "overtime_multiplier": 1.5,
        "total_overtime_payable_hours": 63.0,
        "early_leave_cases": 7,
        "total_early_leave_minutes": 95,
        "early_leave_anomaly_cases": 0,
        "total_break_count": 8,
        "total_break_minutes": 240,
        "employees_with_breaks": 5,
        "incomplete_break_records": 1,
        "employee_id_aliases_used": 24,
        "employee_id_alias_records_mapped": 453,
        "data_quality_score": 98.4,
    }
    return summary, daily


@pytest.fixture
def workbook(tmp_path, monkeypatch):
    monkeypatch.setattr("excel_exporter.REPORT_OUTPUT_DIR", tmp_path)
    summary, daily = _minimal_summary_and_daily()
    export_report(summary, daily,
                  period_start=PERIOD_START, period_end=PERIOD_END)
    written = list(tmp_path.rglob("hr_report_*.xlsx"))
    assert written
    return load_workbook(written[0])


# ---- Zone 1: Header -----------------------------------------------------

def test_dashboard_title_is_large_centered_and_first_row(workbook):
    dash = workbook["Dashboard"]
    cell = dash.cell(row=1, column=1)
    assert cell.value == "HR Reporting Dashboard"
    assert cell.font.bold is True
    assert cell.font.size == 24
    assert cell.alignment.horizontal == "center"


def test_dashboard_gridlines_hidden(workbook):
    assert workbook["Dashboard"].sheet_view.showGridLines is False


def test_dashboard_period_banner_in_rows_2_and_3(workbook):
    dash = workbook["Dashboard"]
    assert "Reporting Period:" in dash.cell(row=2, column=1).value
    assert PERIOD_START in dash.cell(row=2, column=1).value
    assert PERIOD_END in dash.cell(row=2, column=1).value
    assert "Generated On:" in dash.cell(row=3, column=1).value


# ---- Zone 2: 4 hero KPI cards ------------------------------------------

@pytest.mark.parametrize("col_left, section_label, descriptor", [
    ("A", "WORKFORCE",       "Reporting Population"),
    ("D", "ATTENDANCE RISK", "Late Cases"),
    ("G", "PAYROLL IMPACT",  "Total Over Time (Payable 1.5x) (Hours)"),
    ("J", "DATA QUALITY",    "Data Quality Score"),
])
def test_each_hero_card_has_label_value_descriptor(
    workbook, col_left, section_label, descriptor,
):
    """Card spans 5 rows: section label (row 6), value (rows 7-9
    merged), descriptor (row 10). Section label appears UPPERCASE on
    the section's primary accent fill."""
    dash = workbook["Dashboard"]
    top = _DASH_CARDS_ROW
    label_cell = dash[f"{col_left}{top}"]
    assert label_cell.value == section_label
    assert label_cell.font.bold is True
    assert label_cell.font.color.value == "00FFFFFF"  # white

    descriptor_cell = dash[f"{col_left}{top + _DASH_CARD_HEIGHT - 1}"]
    assert descriptor_cell.value == descriptor


def test_hero_card_value_cell_is_large_and_bold(workbook):
    """Card values render at 28pt bold so they read at a glance."""
    dash = workbook["Dashboard"]
    value_cell = dash[f"A{_DASH_CARDS_ROW + 1}"]
    assert value_cell.font.bold is True
    assert value_cell.font.size == 28
    # Workforce card shows reporting_population = 43.
    assert value_cell.value == 43


def test_dashboard_drops_legacy_metric_value_kpi_table(workbook):
    """The pre-Phase-1 dashboard placed `Metric / Value` headers at the
    KPI grid start. Phase 1 replaces that with hero cards + grouped
    sections, so neither label should appear at the legacy positions."""
    dash = workbook["Dashboard"]
    for r in range(3, _DASH_CARDS_ROW):
        assert dash.cell(row=r, column=1).value not in ("Metric",), (
            f"Stray 'Metric' header at row {r}"
        )
        assert dash.cell(row=r, column=2).value not in ("Value",), (
            f"Stray 'Value' header at row {r}"
        )


# ---- Zone 4: Detail KPI sections (A/B/C/D) ------------------------------

def test_detail_section_a_is_workforce_at_expected_row(workbook):
    dash = workbook["Dashboard"]
    cell = dash[f"A{_DASH_DETAIL_SECTIONS_ROW}"]
    assert cell.value == "A. Workforce"
    primary, _ = _DASH_SECTION_COLORS["Workforce"]
    assert cell.fill.fgColor.value.endswith(primary)


def test_detail_section_b_is_attendance_risk_at_expected_row(workbook):
    dash = workbook["Dashboard"]
    # Right detail panel now starts at column H (was I before the
    # A:M fit polish; panel spans H:M, 6 cols, label half H:J).
    cell = dash[f"H{_DASH_DETAIL_SECTIONS_ROW}"]
    assert cell.value == "B. Attendance Risk"
    primary, _ = _DASH_SECTION_COLORS["Attendance Risk"]
    assert cell.fill.fgColor.value.endswith(primary)


def test_detail_data_quality_section_contains_breaks_and_aliases(workbook):
    """Informational KPIs (breaks, aliases) live in 'D. Data Quality
    / Notes' rather than competing for hero-card attention."""
    dash = workbook["Dashboard"]
    # Right detail panel anchors at column H (post-polish). Scan a
    # generous row range so the test survives length tweaks in the
    # other panels.
    found_title = False
    found_break_label = False
    found_alias_label = False
    for r in range(_DASH_DETAIL_SECTIONS_ROW, _DASH_DETAIL_SECTIONS_ROW + 50):
        v = dash.cell(row=r, column=8).value  # right-panel left col (H)
        if v == "D. Data Quality / Notes":
            found_title = True
        if isinstance(v, str) and "Total Break Count" in v:
            found_break_label = True
        if isinstance(v, str) and "Alias Records Mapped" in v:
            found_alias_label = True
    assert found_title, "D. Data Quality / Notes panel not rendered"
    assert found_break_label, "Break KPI missing from Data Quality panel"
    assert found_alias_label, "Alias KPI missing from Data Quality panel"


# ---- Zone 3: Charts pushed below the hero cards -------------------------

def test_charts_anchored_below_hero_cards(workbook):
    """All 5 charts must anchor at or below `_DASH_CHARTS_ROW`. None
    should sit on top of the cards (rows 6-10)."""
    dash = workbook["Dashboard"]
    assert dash._charts, "no charts rendered on the Dashboard"
    cards_bottom = _DASH_CARDS_ROW + _DASH_CARD_HEIGHT - 1
    for chart in dash._charts:
        anchor_row = chart.anchor._from.row + 1  # openpyxl is 0-indexed
        assert anchor_row > cards_bottom, (
            f"chart anchored at row {anchor_row} overlaps cards"
        )
        assert anchor_row >= _DASH_CHARTS_ROW, (
            f"chart anchored at row {anchor_row} below _DASH_CHARTS_ROW"
        )


# ---- Backing tables live on a hidden helper sheet -----------------------

def test_dashboard_data_helper_sheet_exists_and_is_hidden(workbook):
    """The pie + bar charts on the Dashboard reference 3 small tables
    (Attendance Status, Top Overtime Employees, Top Early Leave
    Employees). The polish pass moved those tables off the visible
    Dashboard onto a dedicated `Dashboard Data` helper sheet that is
    flagged `hidden` so HR/payroll users never see it."""
    assert _DASH_HELPER_SHEET_NAME in workbook.sheetnames
    helper = workbook[_DASH_HELPER_SHEET_NAME]
    assert helper.sheet_state == "hidden"
    # Sanity: each of the 3 table headings is present somewhere on it.
    body = {
        helper.cell(row=r, column=1).value
        for r in range(1, helper.max_row + 1)
    }
    assert "Attendance Status" in body
    assert "Top Overtime Employees" in body
    assert "Top Early Leave Employees" in body


def test_visible_dashboard_does_not_carry_underlying_data_heading(workbook):
    """No row of the visible Dashboard should still carry the legacy
    'Underlying Data' section heading -- those tables now live on the
    hidden helper sheet."""
    dash = workbook["Dashboard"]
    for r in range(1, dash.max_row + 1):
        v = dash.cell(row=r, column=1).value
        assert v != "Underlying Data", (
            f"stray 'Underlying Data' heading at Dashboard row {r}"
        )


# ---- Freeze panes pins the top zone -------------------------------------

def test_freeze_pane_keeps_title_banner_and_cards_visible(workbook):
    dash = workbook["Dashboard"]
    # Freeze just below the cards (row 11 with the default geometry).
    expected = f"A{_DASH_CARDS_ROW + _DASH_CARD_HEIGHT}"
    assert dash.freeze_panes == expected


# ---- A:M visible-window invariant (post-polish) -------------------------

def test_all_4_cards_fit_inside_visible_A_M_window(workbook):
    """Every card's right-most column must be at or before column M.
    Previously the Data Quality card extended to column O, which got
    clipped on a laptop viewport."""
    from openpyxl.utils import column_index_from_string
    from excel_exporter import _DASH_CARD_COLS
    M_INDEX = column_index_from_string("M")
    for left, right in _DASH_CARD_COLS:
        assert column_index_from_string(right) <= M_INDEX, (
            f"card {left}:{right} extends past column M"
        )


def test_charts_anchored_within_A_M_window(workbook):
    """Chart anchors must land in cols A:M (1..13). Chart objects extend
    further right than the anchor, but anchoring within the window
    keeps the visual grid tied to the same cols as the card layout."""
    dash = workbook["Dashboard"]
    for chart in dash._charts:
        anchor_col = chart.anchor._from.col + 1  # 1-indexed
        assert 1 <= anchor_col <= 13, (
            f"chart anchored at col {anchor_col} outside A:M"
        )


# ---- Card-to-chart spacing is minimal -----------------------------------

def test_charts_start_immediately_below_cards(workbook):
    """Charts anchor on `_DASH_CHARTS_ROW` (= 11), one row below the
    last card row (10). The pre-polish layout used row 12, leaving
    excessive whitespace between cards and the first chart titles."""
    cards_bottom = _DASH_CARDS_ROW + _DASH_CARD_HEIGHT - 1   # row 10
    assert _DASH_CHARTS_ROW == cards_bottom + 1               # row 11


# ---- Employee Summary First Name column polish --------------------------

def test_employee_summary_first_name_column_widened(workbook):
    """Polish: First Name col (#2) widened to 36 so full HR-style names
    (4+ words ending in -EMPxxx) render without wrapping mid-name."""
    es = workbook["Employee Summary"]
    assert es.column_dimensions["B"].width == 36
